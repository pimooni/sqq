from __future__ import annotations

"""XLSX publication and formatting for shared report tables."""

from time import perf_counter
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import (
    EXCEL_MAX_COLUMNS,
    EXCEL_MAX_ROWS,
    FULL_TABLE_FORMAT_MAX_CELLS,
    FULL_TABLE_FORMAT_MAX_COLUMNS,
    LIGHTWEIGHT_TABLE_COLUMN_WIDTH,
)


def ensure_excel_table_size(
    table: pd.DataFrame,
    sheet_name: str,
    *,
    include_header: bool = True,
) -> None:
    """Fail early with a SQQ-specific message before pandas reaches Excel limits."""
    rows = len(table) + int(include_header)
    columns = len(table.columns)
    if rows > EXCEL_MAX_ROWS or columns > EXCEL_MAX_COLUMNS:
        raise ValueError(
            f"Summary sheet {sheet_name!r} is too large for Excel "
            f"({rows} rows x {columns} columns; limits are "
            f"{EXCEL_MAX_ROWS} rows x {EXCEL_MAX_COLUMNS} columns). "
            "Use summary-detail-csv for high-cardinality detail data."
        )


def format_summary_workbook(workbook) -> list[dict[str, Any]]:
    """Apply formatting and return per-sheet timing and size information."""
    metrics: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        started = perf_counter()
        worksheet.sheet_view.showGridLines = False
        if worksheet.title == "summary":
            format_summary_dashboard_sheet(worksheet)
            format_mode = "dashboard"
        else:
            format_mode = format_table_sheet(worksheet)
        metrics.append(
            {
                "sheet": worksheet.title,
                "rows": int(worksheet.max_row),
                "columns": int(worksheet.max_column),
                "cells": int(worksheet.max_row * worksheet.max_column),
                "format_mode": format_mode,
                "format_seconds": round(perf_counter() - started, 6),
            }
        )
    return metrics


def format_summary_dashboard_sheet(worksheet) -> None:
    """Style the human-facing dashboard sheet."""
    title_fill = PatternFill("solid", fgColor="F7E7C6")
    author_fill = PatternFill("solid", fgColor="FFF7E6")
    section_fill = PatternFill("solid", fgColor="2563EB")
    label_fill = PatternFill("solid", fgColor="EFF6FF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    widths = {"A": 28, "B": 120}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 22
    worksheet.freeze_panes = "A4"

    worksheet.merge_cells("A1:B1")
    worksheet.merge_cells("A2:B2")
    for row in (1, 2):
        for cell in worksheet[row]:
            cell.fill = title_fill if row == 1 else author_fill
            cell.font = Font(color="3B2A14", bold=True, size=18 if row == 1 else 11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    section_labels = {
        "Basic Information",
        "Configuration",
        "Analysis Results (min / mean / max)",
        "Citation Recommendation",
    }
    for row in worksheet.iter_rows():
        if row[0].row <= 2:
            continue
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if cell.value not in (None, ""):
                cell.border = thin_border
        if row[0].value in section_labels:
            for cell in row:
                cell.fill = section_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            if row[0].value not in (None, ""):
                row[0].fill = label_fill
                row[0].font = Font(bold=True, color="0F172A")
            if row[1].value not in (None, ""):
                row[1].font = Font(
                    color="111827",
                    bold=row[0].value == "Recommended text",
                )
                row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def format_table_header(worksheet) -> None:
    """Apply the shared data-sheet header style."""
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def format_table_sheet(worksheet) -> str:
    """Style a data sheet, using lightweight formatting for very large tables."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return "empty"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    format_table_header(worksheet)
    cells = worksheet.max_row * worksheet.max_column
    if cells > FULL_TABLE_FORMAT_MAX_CELLS or worksheet.max_column > FULL_TABLE_FORMAT_MAX_COLUMNS:
        for column_index, cell in enumerate(worksheet[1], start=1):
            letter = get_column_letter(column_index)
            header_width = len(str(cell.value or "")) + 2
            worksheet.column_dimensions[letter].width = min(
                max(10, header_width),
                LIGHTWEIGHT_TABLE_COLUMN_WIDTH,
            )
        return "lightweight"

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        width = estimated_column_width(worksheet, column_index)
        worksheet.column_dimensions[letter].width = width
    no_wrap_columns = {
        column_index
        for column_index, cell in enumerate(worksheet[1], start=1)
        if str(cell.value or "").endswith("_ids")
    }
    for row in worksheet.iter_rows(min_row=2):
        worksheet.row_dimensions[row[0].row].height = 18
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=cell.column not in no_wrap_columns,
            )
    return "full"


def estimated_column_width(worksheet, column_index: int) -> int:
    """Estimate a bounded Excel column width from visible cell values."""
    max_length = 8
    for row_index in range(1, min(worksheet.max_row, 200) + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        max_length = max(max_length, len(str(value)))
    return min(max(max_length + 2, 10), 48)
