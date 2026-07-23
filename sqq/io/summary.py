from __future__ import annotations

"""Markdown, CSV, TSV, VMD, and XLSX summary writers."""

from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime
import os
from pathlib import Path
import re
import tempfile
from time import perf_counter
from typing import Any
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import __version__
from ..banner import SQQ_BANNER
from ..config import (
    dump_config,
    is_cpp_mode,
    mode_display,
    normalize_order_parameters,
    order_parameter_display,
    output_enabled,
    output_type_display,
    q_degrees_from_order_parameters,
)
from ..core.cage import KNOWN_CAGE_TYPES, parse_cage_face_label
from ..display import graph_mode_display
from ..models import CagePatch, FrameResult
from .occupancy import guest_composition_label, guest_lookup as build_guest_lookup, guest_resname_order as guest_resname_order_from_guests


# Avoid per-cell formatting for very large sheets.
FULL_TABLE_FORMAT_MAX_CELLS = 200_000
FULL_TABLE_FORMAT_MAX_COLUMNS = 128
LIGHTWEIGHT_TABLE_COLUMN_WIDTH = 16


# Column order is stable so downstream plotting scripts can rely on it.
SUMMARY_COLUMNS = [
    "frame",
    "time_ps",
    "source",
    "status",
    "error",
    "n_atoms",
    "n_waters",
    "n_guests",
    "bond_mode",
    "n_edges",
    "connection_mode",
    "connection_count",
    "hbond_count",
    "oo_connection_count",
    "pair_connection_count",
    "mean_coordination",
    "coordination_0",
    "coordination_1",
    "coordination_2",
    "coordination_3",
    "coordination_4",
    "coordination_gt4",
    "coordination_0_fraction",
    "coordination_1_fraction",
    "coordination_2_fraction",
    "coordination_3_fraction",
    "coordination_4_fraction",
    "coordination_gt4_fraction",
    "degree_le2_fraction",
    "degree4_fraction",
    "over4_fraction",
    "ring4",
    "ring5",
    "ring6",
    "ring7",
    "free_ring4",
    "free_ring5",
    "free_ring6",
    "free_ring7",
    "half_cage_total",
    "half_cage_breakdown",
    "quasi_cage_total",
    "quasi_cage_breakdown",
    "cage_report_types",
    "cage_512",
    "cage_51262",
    "cage_51263",
    "cage_51264",
    "cage_51268",
    "cage_435663",
    "cage_total",
    "cage_empty",
    "cage_occupied",
    "hydrate_cluster_enabled",
    "hydrate_cluster_detail_enabled",
    "hydrate_cluster_count",
    "sI_cluster_count",
    "sII_cluster_count",
    "sH_cluster_count",
    "mixed_cluster_count",
    "unclassified_cluster_count",
    "hydrate_domain_count",
    "sI_domain_count",
    "sII_domain_count",
    "sH_domain_count",
    "classified_cage_count",
    "boundary_cage_count",
    "ambiguous_cage_count",
    "unclassified_cage_count",
    "isolated_cage_count",
    "largest_cluster_cage_count",
    "largest_cluster_water_count",
    "cluster_size_distribution",
    "MCG1_largest_cluster",
    "DHOP35_largest_cluster",
    "MCG3_largest_cluster",
    "DHOP30_largest_cluster",
    "F3_mean",
    "F4_mean",
    "q6_mean",
    "q12_mean",
    "F3_count",
    "F4_count",
    "q6_count",
    "q12_count",
    "F3_valid_waters",
    "F4_valid_waters",
    "q6_valid_waters",
    "q12_valid_waters",
    "F3_focus_mean",
    "F4_focus_mean",
    "q6_focus_mean",
    "q12_focus_mean",
    "F3_focus_count",
    "F4_focus_count",
    "q6_focus_count",
    "q12_focus_count",
    "F3_focus_valid_waters",
    "F4_focus_valid_waters",
    "q6_focus_valid_waters",
    "q12_focus_valid_waters",
    "ice_like_waters",
    "ice_i_waters",
    "interfacial_ice_waters",
]

SUMMARY_DETAIL_TABLE_NAMES = (
    "failures",
    "cage_occupancy",
    "cage_isomer",
    "quasi_cage_isomer",
    "hydrate_domain",
    "hydrate_cluster_detail",
)

# Main summary CSV filenames mirror the workbook sheet names. Keeping the
# allow-list explicit prevents cleanup from touching unrelated files.
SUMMARY_MAIN_TABLE_NAMES = (
    "summary",
    "failures",
    "connection",
    "hbond",
    "oo_connection",
    "pair_connection",
    "ring",
    "half_cage",
    "quasi_cage",
    "cage",
    "cage_occupancy",
    "cage_isomer",
    "hydrate_cluster",
    "order_parameter",
    "ice",
    "detail_index",
    "config",
)

TREE_MIDDLE = "\u251c"
TREE_LAST = "\u2514"
TREE_PIPE = "\u2502"
SUBSCRIPT_DIGIT_DELETE = dict.fromkeys(range(0x2080, 0x208A))
QUASI_ISOMER_DETAIL_KEY = "_quasi_cage_isomer_detail"
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384


def result_row(
    result: FrameResult,
    *,
    include_cluster_details: bool = True,
) -> dict[str, Any]:
    """Flatten a FrameResult into one summary-table row."""
    cage_counts: dict[str, int] = {}
    cage_detail_counts: dict[str, int] = {}
    cage_isomers: dict[str, dict[str, int]] = {}
    guests_by_id = build_guest_lookup(result.guests)
    guest_order = guest_resname_order(result)
    molecule_counts = atom_resname_counts(result)

    for cage in result.cages:
        cage_counts[cage.cage_type] = cage_counts.get(cage.cage_type, 0) + 1
        occupancy_key = f"cage_{cage.cage_type}_{'occupied' if cage.occupied else 'empty'}"
        cage_detail_counts[occupancy_key] = cage_detail_counts.get(occupancy_key, 0) + 1
        isomer = cage.isomer or "plain"
        type_isomers = cage_isomers.setdefault(cage.cage_type, {})
        type_isomers[isomer] = type_isomers.get(isomer, 0) + 1
        if cage.occupied:
            composition = guest_composition_label(cage, guests_by_id, guest_order)
            if composition:
                guest_key = f"cage_{cage.cage_type}_{composition}"
                cage_detail_counts[guest_key] = cage_detail_counts.get(guest_key, 0) + 1
            if len(cage.guest_ids) > 1:
                multi_key = f"cage_{cage.cage_type}_multi"
                cage_detail_counts[multi_key] = cage_detail_counts.get(multi_key, 0) + 1

    empty = sum(1 for cage in result.cages if not cage.occupied)
    occupied = sum(1 for cage in result.cages if cage.occupied)
    used_ring_ids = {ring_id for patch in [*result.half_cages, *result.quasi_cages] for ring_id in patch.rings}
    filtering_cages = result.all_cages or result.cages
    used_ring_ids.update(ring_id for cage in filtering_cages for ring_id in cage.rings)
    half_cage_counts = patch_counts(result.half_cages)
    quasi_cage_counts = patch_counts(result.quasi_cages)
    quasi_composition_counts: dict[str, int] = {}
    for patch_type, count in quasi_cage_counts.items():
        composition = patch_composition_label(patch_type)
        quasi_composition_counts[composition] = quasi_composition_counts.get(composition, 0) + count
    cluster_details = (
        hydrate_cluster_detail_records(result, guests_by_id, guest_order)
        if include_cluster_details and result.hydrate_cluster_detail
        else []
    )
    domain_details = (
        hydrate_domain_records(result, guests_by_id, guest_order)
        if include_cluster_details and result.hydrate_cluster_detail
        else []
    )
    largest_cluster = max(result.hydrate_clusters, key=lambda cluster: cluster.cage_count, default=None)
    cluster_type_counts = Counter(cluster.hydrate_type for cluster in result.hydrate_clusters)
    domain_type_counts = Counter(domain.hydrate_type for domain in result.hydrate_domains)
    classified_cage_count = sum(
        len(cluster.classified_cage_ids) for cluster in result.hydrate_clusters
    )
    boundary_cage_count = sum(
        len(cluster.boundary_cage_ids) for cluster in result.hydrate_clusters
    )
    ambiguous_cage_count = sum(
        len(cluster.ambiguous_cage_ids) for cluster in result.hydrate_clusters
    )
    unclassified_cage_count = sum(
        len(cluster.unclassified_cage_ids) for cluster in result.hydrate_clusters
    )

    def free_count(size: int) -> int:
        # Free rings are rings not consumed by open patches or cages.
        return sum(1 for ring in result.rings.get(size, []) if ring.object_id not in used_ring_ids)

    f3f4 = result.f3f4
    connection_counts = graph_connection_counts(result)
    row: dict[str, Any] = {
        "frame": result.frame.name,
        "time_ps": result.frame.time_ps,
        "source": source_label(result.frame.source),
        "status": "ok",
        "error": "",
        "n_atoms": len(result.frame.atoms),
        "n_waters": len(result.waters),
        "n_guests": len(result.guests),
        "bond_mode": result.graph.mode,
        "n_edges": len(result.graph.edges),
        **connection_counts,
        "ring4": len(result.rings.get(4, [])),
        "ring5": len(result.rings.get(5, [])),
        "ring6": len(result.rings.get(6, [])),
        "ring7": len(result.rings.get(7, [])),
        "free_ring4": free_count(4),
        "free_ring5": free_count(5),
        "free_ring6": free_count(6),
        "free_ring7": free_count(7),
        "half_cage_total": len(result.half_cages),
        "half_cage_breakdown": patch_breakdown(half_cage_counts),
        "quasi_cage_total": len(result.quasi_cages),
        "quasi_cage_breakdown": patch_breakdown(quasi_cage_counts),
        "cage_report_types": "all" if result.cage_report_types is None else ";".join(result.cage_report_types),
        "cage_512": cage_counts.get("512", 0),
        "cage_51262": cage_counts.get("51262", 0),
        "cage_51263": cage_counts.get("51263", 0),
        "cage_51264": cage_counts.get("51264", 0),
        "cage_51268": cage_counts.get("51268", 0),
        "cage_435663": cage_counts.get("435663", 0),
        "cage_total": len(result.cages),
        "cage_empty": empty,
        "cage_occupied": occupied,
        "hydrate_cluster_enabled": "on" if result.hydrate_cluster_enabled else "off",
        "hydrate_cluster_detail_enabled": "on" if result.hydrate_cluster_detail else "off",
        "hydrate_cluster_count": len(result.hydrate_clusters) if result.hydrate_cluster_enabled else 0,
        "sI_cluster_count": cluster_type_counts.get("sI", 0),
        "sII_cluster_count": cluster_type_counts.get("sII", 0),
        "sH_cluster_count": cluster_type_counts.get("sH", 0),
        "mixed_cluster_count": cluster_type_counts.get("mixed", 0),
        "unclassified_cluster_count": cluster_type_counts.get("unclassified", 0),
        "hydrate_domain_count": len(result.hydrate_domains),
        "sI_domain_count": domain_type_counts.get("sI", 0),
        "sII_domain_count": domain_type_counts.get("sII", 0),
        "sH_domain_count": domain_type_counts.get("sH", 0),
        "classified_cage_count": classified_cage_count,
        "boundary_cage_count": boundary_cage_count,
        "ambiguous_cage_count": ambiguous_cage_count,
        "unclassified_cage_count": unclassified_cage_count,
        "isolated_cage_count": len(result.isolated_cage_ids) if result.hydrate_cluster_enabled else 0,
        "largest_cluster_cage_count": 0 if largest_cluster is None else largest_cluster.cage_count,
        "largest_cluster_water_count": 0 if largest_cluster is None else largest_cluster.water_count,
        "cluster_size_distribution": cluster_size_distribution(result.hydrate_clusters),
        "MCG1_largest_cluster": None if result.hydrate_order is None else result.hydrate_order.mcg1.largest_cluster_size,
        "DHOP35_largest_cluster": None if result.hydrate_order is None else result.hydrate_order.dhop35.largest_cluster_size,
        "MCG3_largest_cluster": None if result.hydrate_order is None or result.hydrate_order.mcg3 is None else result.hydrate_order.mcg3.largest_cluster_size,
        "DHOP30_largest_cluster": None if result.hydrate_order is None or result.hydrate_order.dhop30 is None else result.hydrate_order.dhop30.largest_cluster_size,
        "MCG3_enabled": result.hydrate_order is not None and result.hydrate_order.mcg3 is not None,
        "DHOP30_enabled": result.hydrate_order is not None and result.hydrate_order.dhop30 is not None,
        "hydrate_cluster_detail": cluster_details,
        "hydrate_domain_detail": domain_details,
        "F3_mean": None if f3f4 is None else f3f4.f3_mean,
        "F4_mean": None if f3f4 is None else f3f4.f4_mean,
        "F3_count": None if f3f4 is None else f3f4.f3_valid,
        "F4_count": None if f3f4 is None else f3f4.f4_valid,
        "F3_valid_waters": None if f3f4 is None else f3f4.f3_valid,
        "F4_valid_waters": None if f3f4 is None else f3f4.f4_valid,
        "F3_focus_mean": None if f3f4 is None else f3f4.f3_focus_mean,
        "F4_focus_mean": None if f3f4 is None else f3f4.f4_focus_mean,
        "F3_focus_count": None if f3f4 is None else f3f4.f3_focus_valid,
        "F4_focus_count": None if f3f4 is None else f3f4.f4_focus_valid,
        "F3_focus_valid_waters": None if f3f4 is None else f3f4.f3_focus_valid,
        "F4_focus_valid_waters": None if f3f4 is None else f3f4.f4_focus_valid,
        "ice_like_waters": len(result.ice_like_waters),
        "ice_i_waters": len(result.ice_i_waters),
        "interfacial_ice_waters": len(result.interfacial_ice_waters),
        # Keep exact quasi isomers out of the wide summary table.
        QUASI_ISOMER_DETAIL_KEY: tuple(sorted(quasi_cage_counts.items())),
    }
    if f3f4 is not None:
        for degree in f3f4.q_degree:
            prefix = f"q{degree}"
            row[f"{prefix}_mean"] = f3f4.q_means.get(degree)
            row[f"{prefix}_count"] = f3f4.q_valid_counts.get(degree, 0)
            row[f"{prefix}_valid_waters"] = f3f4.q_valid_counts.get(degree, 0)
            row[f"{prefix}_focus_mean"] = f3f4.q_focus_means.get(degree)
            row[f"{prefix}_focus_count"] = f3f4.q_focus_valid_counts.get(degree, 0)
            row[f"{prefix}_focus_valid_waters"] = f3f4.q_focus_valid_counts.get(degree, 0)
    else:
        for degree in (6, 12):
            prefix = f"q{degree}"
            row[f"{prefix}_mean"] = None
            row[f"{prefix}_count"] = None
            row[f"{prefix}_valid_waters"] = None
            row[f"{prefix}_focus_mean"] = None
            row[f"{prefix}_focus_count"] = None
            row[f"{prefix}_focus_valid_waters"] = None
    for resname, count in molecule_counts.items():
        row[f"mol_{resname}"] = count
    row["mol_TOTAL"] = len(result.frame.atoms)
    row["guest_order"] = ";".join(guest_order)
    for patch_type, count in half_cage_counts.items():
        row[f"half_cage_{patch_type}"] = count
    for composition, count in quasi_composition_counts.items():
        row[f"quasi_cage_{composition}"] = count

    cage_types = ordered_cage_types(cage_counts)
    for cage_type in cage_types:
        prefix = f"cage_{cage_type}"
        row[prefix] = cage_counts.get(cage_type, row.get(prefix, 0))
        row[f"{prefix}_empty"] = cage_detail_counts.get(f"{prefix}_empty", 0)
        row[f"{prefix}_occupied"] = cage_detail_counts.get(f"{prefix}_occupied", 0)
        row[f"{prefix}_multi"] = cage_detail_counts.get(f"{prefix}_multi", 0)
        if cage_type in cage_isomers:
            parts = [f"{key}:{cage_isomers[cage_type][key]}" for key in sorted(cage_isomers[cage_type])]
            row[f"{prefix}_isomers"] = "; ".join(parts)
            for isomer, count in cage_isomers[cage_type].items():
                row[f"{prefix}_isomer_{isomer}"] = count

    for key in sorted(cage_detail_counts):
        if key not in row:
            row[key] = cage_detail_counts[key]
    return row


def graph_connection_counts(result: FrameResult) -> dict[str, Any]:
    """Return graph counts and a diagnostic-only coordination distribution."""
    mode = result.graph.mode
    edge_count = len(result.graph.edges)
    degrees = [len(result.graph.adjacency.get(water.oxygen, set())) for water in result.waters]
    water_count = len(degrees)
    bins = {degree: sum(value == degree for value in degrees) for degree in range(5)}
    over_four = sum(value > 4 for value in degrees)

    def fraction(count: int) -> float:
        return 0.0 if water_count == 0 else count / water_count

    values: dict[str, Any] = {
        "connection_mode": mode,
        "connection_count": edge_count,
        "hbond_count": edge_count if mode == "hbond" else None,
        "oo_connection_count": edge_count if mode == "oo" else None,
        "pair_connection_count": edge_count if mode == "pairs" else None,
        "mean_coordination": 0.0 if water_count == 0 else sum(degrees) / water_count,
        "coordination_gt4": over_four,
        "coordination_gt4_fraction": fraction(over_four),
        "degree_le2_fraction": fraction(sum(bins[degree] for degree in range(3))),
        "degree4_fraction": fraction(bins[4]),
        "over4_fraction": fraction(over_four),
    }
    for degree in range(5):
        values[f"coordination_{degree}"] = bins[degree]
        values[f"coordination_{degree}_fraction"] = fraction(bins[degree])
    return values


def hydrate_cluster_info_section(
    result: FrameResult,
    row: dict[str, Any] | None = None,
) -> list[str]:
    """Render the compact, mutually exclusive cluster hierarchy."""
    if not result.hydrate_cluster_enabled:
        return []

    cage_by_id = {
        cage.object_id: cage for cage in (result.all_cages or result.cages)
    }
    domains_by_cluster: dict[str, list[Any]] = defaultdict(list)
    for domain in result.hydrate_domains:
        domains_by_cluster[domain.cluster_id].append(domain)

    rows: list[list[Any]] = []
    displayed_cluster_ids: set[str] = set()
    for cluster in sorted(result.hydrate_clusters, key=lambda item: item.object_id):
        cluster_ids = tuple(dict.fromkeys(cluster.cage_ids))
        cluster_id_set = set(cluster_ids)
        displayed_cluster_ids.update(cluster_id_set)
        remaining_ids = set(cluster_ids)
        rows.append(
            [cluster.object_id, cluster.hydrate_type, len(cluster_id_set)]
        )

        children: list[tuple[str, str, tuple[str, ...]]] = []
        domains = sorted(
            domains_by_cluster.get(cluster.object_id, []),
            key=lambda item: item.object_id,
        )
        for domain in domains:
            domain_ids = tuple(
                cage_id
                for cage_id in dict.fromkeys(domain.cage_ids)
                if cage_id in remaining_ids
            )
            if not domain_ids:
                continue
            remaining_ids.difference_update(domain_ids)
            children.append(
                (domain.object_id, domain.hydrate_type, domain_ids)
            )

        boundary_ids = tuple(
            cage_id
            for cage_id in dict.fromkeys(cluster.boundary_cage_ids)
            if cage_id in remaining_ids
        )
        if boundary_ids:
            remaining_ids.difference_update(boundary_ids)
            children.append(("boundary", "boundary", boundary_ids))

        unclassified_ids = tuple(
            cage_id for cage_id in cluster_ids if cage_id in remaining_ids
        )
        if unclassified_ids:
            children.append(
                ("unclassified", "unclassified", unclassified_ids)
            )

        for child_index, (name, hydrate_type, cage_ids) in enumerate(children):
            child_branch = (
                TREE_LAST if child_index == len(children) - 1 else TREE_MIDDLE
            )
            rows.append(
                [
                    f"{child_branch} {name}",
                    hydrate_type,
                    f"{child_branch} {len(cage_ids)}",
                ]
            )
            type_counts = Counter(
                cage_by_id[cage_id].cage_type
                for cage_id in cage_ids
                if cage_id in cage_by_id
            )
            cage_types = present_cage_types(type_counts)
            for type_index, cage_type in enumerate(cage_types):
                type_branch = (
                    TREE_LAST if type_index == len(cage_types) - 1 else TREE_MIDDLE
                )
                rows.append(
                    [
                        f"  {type_branch} {cage_display_label(cage_type)}",
                        "",
                        f"  {type_branch} {type_counts[cage_type]}",
                    ]
                )

    isolated_count = len(
        set(result.isolated_cage_ids).difference(displayed_cluster_ids)
    )
    if isolated_count:
        rows.append(["isolated", "isolated", isolated_count])

    if not rows:
        return ["", "## Hydrate Cluster", "", "no hydrate cluster"]
    return section_table(
        "Hydrate Cluster",
        ["item", "type", "cage_qty"],
        rows,
    )


def hydrate_cluster_detail_section(records: list[dict[str, Any]]) -> list[str]:
    """Render one vertical table per cluster, including cage composition."""
    if not records:
        return section_table(
            "Hydrate Cluster Detail",
            ["item", "value"],
            [["cluster_count", 0]],
        )
    lines = ["", "## Hydrate Cluster Detail", ""]
    metric_keys = (
        "hydrate_type",
        "cage_count",
        "classified_cage_count",
        "boundary_cage_count",
        "ambiguous_cage_count",
        "unclassified_cage_count",
        "classified_cage_fraction",
        "domain_count",
        "water_count",
        "guest_count",
        "empty_cage_count",
        "occupied_cage_count",
        "boundary_composition",
        "guest_composition",
    )
    for record in records:
        rows = [[key, record.get(key, "")] for key in metric_keys]
        type_counts = record.get("cage_type_counts", {})
        if type_counts:
            rows.append(["cage composition", ""])
            ordered_types = present_cage_types(type_counts)
            for index, cage_type in enumerate(ordered_types):
                branch = TREE_LAST if index == len(ordered_types) - 1 else TREE_MIDDLE
                rows.append(
                    [f"{branch} {cage_display_label(cage_type)}", type_counts[cage_type]]
                )
        lines.extend(object_vertical_table(str(record["cluster_id"]), rows))
    return lines


def hydrate_cluster_hierarchy_section(result: FrameResult) -> list[str]:
    """Render mutually exclusive cluster categories."""
    cage_by_id = {cage.object_id: cage for cage in (result.all_cages or result.cages)}
    domains_by_cluster: dict[str, list[Any]] = defaultdict(list)
    for domain in result.hydrate_domains:
        domains_by_cluster[domain.cluster_id].append(domain)
    if not result.hydrate_clusters:
        return ["", "## Hydrate Cluster Hierarchy", "", "no hydrate cluster"]

    lines = ["", "## Hydrate Cluster Hierarchy", ""]
    for cluster_index, cluster in enumerate(result.hydrate_clusters):
        if cluster_index:
            lines.append("")
        cluster_counts = Counter(
            cage_by_id[cage_id].cage_type
            for cage_id in cluster.cage_ids
            if cage_id in cage_by_id
        )
        cage_types = present_cage_types(cluster_counts)
        headers = [
            "item",
            "type",
            "cage_qty",
            *(cage_display_label(cage_type) for cage_type in cage_types),
        ]
        rows: list[list[Any]] = [
            hierarchy_table_row(
                cluster.object_id,
                cluster.hydrate_type,
                cluster.cage_ids,
                cage_types,
                cage_by_id,
            )
        ]
        children: list[tuple[str, str, tuple[str, ...]]] = [
            (domain.object_id, domain.hydrate_type, domain.cage_ids)
            for domain in domains_by_cluster.get(cluster.object_id, [])
        ]
        if cluster.boundary_cage_ids:
            children.append(("boundary", "boundary", cluster.boundary_cage_ids))
        if cluster.ambiguous_cage_ids:
            children.append(("ambiguous", "ambiguous", cluster.ambiguous_cage_ids))
        if cluster.unclassified_cage_ids:
            children.append(
                ("unclassified", "unclassified", cluster.unclassified_cage_ids)
            )
        for child_index, (name, hydrate_type, cage_ids) in enumerate(children):
            branch = TREE_LAST if child_index == len(children) - 1 else TREE_MIDDLE
            rows.append(
                hierarchy_table_row(
                    f"{branch} {name}",
                    hydrate_type,
                    cage_ids,
                    cage_types,
                    cage_by_id,
                )
            )
        lines.append(markdown_rows(headers, rows).rstrip())
    return lines


def hierarchy_table_row(
    item: str,
    hydrate_type: str,
    cage_ids: tuple[str, ...],
    cage_types: list[str],
    cage_by_id: dict[str, Any],
) -> list[Any]:
    """Return one hierarchy row; absent cage types are rendered as dashes."""
    counts = Counter(
        cage_by_id[cage_id].cage_type
        for cage_id in cage_ids
        if cage_id in cage_by_id
    )
    return [
        item,
        hydrate_type,
        len(cage_ids),
        *(counts[cage_type] if counts[cage_type] else "-" for cage_type in cage_types),
    ]


def hierarchy_label(
    object_id: str,
    hydrate_type: str,
    cage_ids: tuple[str, ...],
    cage_by_id: dict[str, Any],
) -> str:
    """Return one readable tree label with nonzero cage-type counts."""
    counts = Counter(cage_by_id[cage_id].cage_type for cage_id in cage_ids if cage_id in cage_by_id)
    parts = [hydrate_type] if hydrate_type else []
    parts.append(f"cages={len(cage_ids)}")
    parts.extend(f"{cage_display_label(cage_type)}={counts[cage_type]}" for cage_type in present_cage_types(counts))
    return f"{object_id} [{', '.join(parts)}]"


def domain_hierarchy_label(domain: Any, cage_by_id: dict[str, Any]) -> str:
    """Return a domain label with internal seed and expansion counts."""
    seed_ids = set(domain.seed_cage_ids)
    counts = Counter(cage_by_id[cage_id].cage_type for cage_id in domain.cage_ids if cage_id in cage_by_id)
    parts = [
        domain.hydrate_type,
        domain.status,
        f"cages={domain.cage_count}",
        f"seeds={domain.seed_count}",
        f"seed_cages={len(seed_ids)}",
        f"expanded={domain.cage_count - len(seed_ids)}",
    ]
    parts.extend(f"{cage_display_label(cage_type)}={counts[cage_type]}" for cage_type in present_cage_types(counts))
    return f"{domain.object_id} [{', '.join(parts)}]"


def motif_hierarchy_label(motif: Any, cage_by_id: dict[str, Any]) -> str:
    """Return a compact label for one overlapping local topology motif."""
    counts = Counter(cage_by_id[cage_id].cage_type for cage_id in motif.cage_ids if cage_id in cage_by_id)
    core_count = len(motif.anchor_cage_ids)
    parts = [
        motif.hydrate_type,
        motif.status,
        f"cages={motif.cage_count}",
        f"completeness={motif.completeness:.2f}",
        f"core={core_count}",
        f"support={motif.cage_count - core_count}",
    ]
    parts.extend(f"{cage_display_label(cage_type)}={counts[cage_type]}" for cage_type in present_cage_types(counts))
    return f"{motif.object_id} [{', '.join(parts)}]"


def object_vertical_table(object_id: str, rows: list[list[Any]]) -> list[str]:
    """Render one level-three object heading followed by a narrow table."""
    return [f"### {object_id}", "", markdown_rows(["item", object_id], rows).rstrip(), ""]


def hydrate_domain_info_section(
    result: FrameResult,
    guests_by_id: dict[str, Any],
    guest_order: list[str],
) -> list[str]:
    """Render one vertical detail table per hydrate domain."""
    records = hydrate_domain_records(result, guests_by_id, guest_order)
    if not records:
        return section_table("Hydrate Domain", ["item", "value"], [["domain_count", 0]])
    lines = ["", "## Hydrate Domain", ""]
    keys = (
        "cluster_id",
        "hydrate_type",
        "status",
        "cage_count",
        "seed_count",
        "seed_cage_count",
        "expanded_cage_count",
        "classified_fraction",
        "water_count",
        "guest_count",
        "external_boundary_contact_count",
        "cage_composition",
        "guest_composition",
    )
    for record in records:
        rows = [[key, record.get(key, "")] for key in keys]
        lines.extend(object_vertical_table(str(record["domain_id"]), rows))
    return lines


def hydrate_motif_info_section(result: FrameResult) -> list[str]:
    """Render one vertical evidence table per hydrate motif."""
    records = hydrate_motif_records(result)
    if not records:
        return section_table("Hydrate Motif", ["item", "value"], [["motif_count", 0]])
    lines = ["", "## Hydrate Motif", ""]
    keys = (
        "cluster_id",
        "domain_id",
        "hydrate_type",
        "status",
        "completeness",
        "consistency",
        "confidence",
        "cage_count",
        "core_cage_count",
        "support_cage_count",
        "cage_composition",
        "core_cage_composition",
        "internal_shared_face_count",
        "classification_method",
    )
    for record in records:
        rows = [[key, record.get(key, "")] for key in keys]
        lines.extend(object_vertical_table(str(record["motif_id"]), rows))
    return lines


def hydrate_boundary_info_section(result: FrameResult) -> list[str]:
    """Render exclusive boundary totals and composition."""
    clusters = [
        cluster for cluster in result.hydrate_clusters if cluster.boundary_cage_ids
    ]
    if not clusters:
        return section_table(
            "Hydrate Boundary",
            ["item", "value"],
            [["boundary_cage_count", 0]],
        )
    lines = ["", "## Hydrate Boundary", ""]
    cage_by_id = {cage.object_id: cage for cage in (result.all_cages or result.cages)}
    for cluster in clusters:
        counts = Counter(
            cage_by_id[cage_id].cage_type
            for cage_id in cluster.boundary_cage_ids
            if cage_id in cage_by_id
        )
        rows = [
            ["boundary_cage_count", len(cluster.boundary_cage_ids)],
            ["boundary_composition", format_cage_type_counts(counts)],
        ]
        lines.extend(object_vertical_table(cluster.object_id, rows))
    return lines


def format_cage_type_counts(counts: Counter) -> str:
    """Format nonzero cage-type counts for compact info tables."""
    return ";".join(
        f"{cage_display_label(cage_type)}:{counts[cage_type]}"
        for cage_type in present_cage_types(counts)
    )


def append_tree_value_rows(rows: list[list[Any]], label: str, item_label: str, values: list[str]) -> None:
    """Append a vertical, consistently branched list to a two-column table."""
    rows.append([label, ""])
    for index, value in enumerate(values):
        branch = TREE_LAST if index == len(values) - 1 else TREE_MIDDLE
        rows.append([f"{branch} {item_label}", value])


def split_record_ids(value: Any) -> list[str]:
    """Split a semicolon-delimited workbook field for vertical info output."""
    return [item for item in str(value).split(";") if item]


def hydrate_cluster_detail_records(
    result: FrameResult,
    guests_by_id: dict[str, Any],
    guest_order: list[str],
) -> list[dict[str, Any]]:
    """Return one plotting-friendly detail record per hydrate cluster."""
    if not result.hydrate_cluster_enabled:
        return []
    cage_by_id = {cage.object_id: cage for cage in (result.all_cages or result.cages)}
    records: list[dict[str, Any]] = []
    for cluster in result.hydrate_clusters:
        cluster_cages = [
            cage_by_id[cage_id]
            for cage_id in cluster.cage_ids
            if cage_id in cage_by_id
        ]
        boundary_cages = [
            cage_by_id[cage_id]
            for cage_id in cluster.boundary_cage_ids
            if cage_id in cage_by_id
        ]
        type_counts = Counter(cage.cage_type for cage in cluster_cages)
        boundary_counts = Counter(cage.cage_type for cage in boundary_cages)
        ordered_types = present_cage_types(type_counts)
        classified_fraction = (
            0.0
            if not cluster.cage_count
            else len(cluster.classified_cage_ids) / cluster.cage_count
        )
        records.append(
            {
                "cluster_id": cluster.object_id,
                "hydrate_type": cluster.hydrate_type,
                "cage_count": cluster.cage_count,
                "water_count": cluster.water_count,
                "guest_count": cluster.guest_count,
                "empty_cage_count": sum(
                    1 for cage in cluster_cages if not cage.occupied
                ),
                "occupied_cage_count": sum(
                    1 for cage in cluster_cages if cage.occupied
                ),
                "classified_cage_count": len(cluster.classified_cage_ids),
                "boundary_cage_count": len(cluster.boundary_cage_ids),
                "ambiguous_cage_count": len(cluster.ambiguous_cage_ids),
                "unclassified_cage_count": len(cluster.unclassified_cage_ids),
                "classified_cage_fraction": classified_fraction,
                "domain_count": cluster.domain_count,
                "cage_type_counts": {
                    cage_type: type_counts[cage_type] for cage_type in ordered_types
                },
                "cage_composition": ";".join(
                    f"{cage_display_label(cage_type)}:{type_counts[cage_type]}"
                    for cage_type in ordered_types
                ),
                "boundary_composition": format_cage_type_counts(boundary_counts),
                "guest_composition": cluster_guest_composition(
                    cluster.guest_ids,
                    guests_by_id,
                    guest_order,
                ),
                "domain_ids": ";".join(cluster.domain_ids),
                "cage_ids": ";".join(cluster.cage_ids),
                "classified_cage_ids": ";".join(cluster.classified_cage_ids),
                "boundary_cage_ids": ";".join(cluster.boundary_cage_ids),
                "ambiguous_cage_ids": ";".join(cluster.ambiguous_cage_ids),
                "unclassified_cage_ids": ";".join(cluster.unclassified_cage_ids),
                "shared_face_count": len(cluster.shared_faces),
            }
        )
    return records


def hydrate_domain_records(
    result: FrameResult,
    guests_by_id: dict[str, Any],
    guest_order: list[str],
) -> list[dict[str, Any]]:
    """Return one plotting-friendly record per hydrate domain."""
    if not result.hydrate_cluster_enabled:
        return []
    cage_by_id = {cage.object_id: cage for cage in (result.all_cages or result.cages)}
    records: list[dict[str, Any]] = []
    for domain in result.hydrate_domains:
        domain_cages = [cage_by_id[cage_id] for cage_id in domain.cage_ids if cage_id in cage_by_id]
        type_counts = Counter(cage.cage_type for cage in domain_cages)
        ordered_types = present_cage_types(type_counts)
        seed_cage_ids = set(domain.seed_cage_ids)
        records.append(
            {
                "domain_id": domain.object_id,
                "cluster_id": domain.cluster_id,
                "hydrate_type": domain.hydrate_type,
                "status": domain.status,
                "cage_count": domain.cage_count,
                "seed_count": domain.seed_count,
                "seed_cage_count": len(seed_cage_ids),
                "expanded_cage_count": domain.cage_count - len(seed_cage_ids),
                "classified_fraction": domain.classified_fraction,
                "water_count": domain.water_count,
                "guest_count": domain.guest_count,
                "external_boundary_contact_count": len(domain.boundary_cage_ids),
                "cage_composition": ";".join(f"{cage_display_label(cage_type)}:{type_counts[cage_type]}" for cage_type in ordered_types),
                "guest_composition": cluster_guest_composition(domain.guest_ids, guests_by_id, guest_order),
                "cage_ids": ";".join(domain.cage_ids),
                "seed_cage_ids": ";".join(domain.seed_cage_ids),
                "external_boundary_contact_ids": ";".join(domain.boundary_cage_ids),
            }
        )
    return records


def hydrate_motif_records(result: FrameResult) -> list[dict[str, Any]]:
    """Return one record per overlapping local topology motif."""
    if not result.hydrate_cluster_enabled:
        return []
    cage_by_id = {cage.object_id: cage for cage in (result.all_cages or result.cages)}
    records: list[dict[str, Any]] = []
    for motif in result.hydrate_motifs:
        member_cages = [cage_by_id[cage_id] for cage_id in motif.cage_ids if cage_id in cage_by_id]
        type_counts = Counter(cage.cage_type for cage in member_cages)
        ordered_types = present_cage_types(type_counts)
        core_cages = [cage_by_id[cage_id] for cage_id in motif.anchor_cage_ids if cage_id in cage_by_id]
        core_type_counts = Counter(cage.cage_type for cage in core_cages)
        ordered_core_types = present_cage_types(core_type_counts)
        anchor_types = [cage_display_label(cage_by_id[cage_id].cage_type) for cage_id in motif.anchor_cage_ids if cage_id in cage_by_id]
        records.append(
            {
                "motif_id": motif.object_id,
                "cluster_id": motif.cluster_id,
                "domain_id": motif.domain_id,
                "hydrate_type": motif.hydrate_type,
                "status": motif.status,
                "completeness": motif.completeness,
                "consistency": motif.consistency,
                "confidence": motif.confidence,
                "anchor_cage_types": ";".join(anchor_types),
                "anchor_cage_ids": ";".join(motif.anchor_cage_ids),
                "member_cage_count": motif.cage_count,
                "cage_count": motif.cage_count,
                "support_cage_count": motif.cage_count - len(motif.anchor_cage_ids),
                "cage_composition": ";".join(f"{cage_display_label(cage_type)}:{type_counts[cage_type]}" for cage_type in ordered_types),
                "core_cage_count": len(motif.anchor_cage_ids),
                "core_cage_composition": ";".join(f"{cage_display_label(cage_type)}:{core_type_counts[cage_type]}" for cage_type in ordered_core_types),
                "core_cage_ids": ";".join(motif.anchor_cage_ids),
                "motif_cage_ids": ";".join(motif.cage_ids),
                "member_cage_ids": ";".join(motif.cage_ids),
                "shared_face_count": len(motif.shared_face_ids),
                "internal_shared_face_count": len(motif.shared_face_ids),
                "shared_face_ids": ";".join(motif.shared_face_ids),
                "internal_shared_face_ids": ";".join(motif.shared_face_ids),
                "classification_method": motif.classification_method,
            }
        )
    return records


def cluster_guest_composition(guest_ids: tuple[str, ...], guests_by_id: dict[str, Any], guest_order: list[str]) -> str:
    """Summarize all guest residue names inside one hydrate cluster."""
    names = [guests_by_id[item].resname for item in guest_ids if item in guests_by_id]
    if not names:
        return ""
    counts = Counter(names)
    order_index = {name: index for index, name in enumerate(guest_order)}
    ordered_names = sorted(counts, key=lambda name: (order_index.get(name, 10_000), name))
    return "+".join(name if counts[name] == 1 else f"{name}x{counts[name]}" for name in ordered_names)


def cluster_size_distribution(clusters) -> str:
    """Summarize cluster sizes as cage_count:number_of_clusters."""
    if not clusters:
        return ""
    counts = Counter(cluster.cage_count for cluster in clusters)
    return ";".join(f"{size}:{counts[size]}" for size in sorted(counts))


def patch_counts(patches) -> dict[str, int]:
    """Count open cage patches by patch_type."""
    counts: dict[str, int] = {}
    for patch in patches:
        counts[patch.patch_type] = counts.get(patch.patch_type, 0) + 1
    return counts


def patch_breakdown(counts: dict[str, int]) -> str:
    """Render a compact patch count list for broad summary rows."""
    return "; ".join(f"{key}:{counts[key]}" for key in sorted(counts))


def failed_row(frame_name: str, source: str, error: str) -> dict[str, Any]:
    """Create a summary row for a skipped or failed frame."""
    row = {column: "" for column in SUMMARY_COLUMNS}
    row.update({"frame": frame_name, "source": source, "status": "failed", "error": error})
    return row


def write_frame_info(
    result: FrameResult,
    frame_dir: Path,
    ring_sizes: list[int] | None = None,
    requested_bond_mode: Any | None = None,
    order_parameters: Any | None = None,
    analysis_mode: Any = "50",
    input_metadata: dict[str, Any] | None = None,
) -> None:
    """Write the per-frame Markdown report with inspection-oriented tables."""
    frame_dir.mkdir(parents=True, exist_ok=True)
    row = result_row(result, include_cluster_details=False)
    selected_order_parameters = selected_order_parameters_for_result(
        result,
        order_parameters,
    )
    selected_order_set = set(selected_order_parameters)
    cage_values = {cage.cage_type for cage in result.cages}
    cage_types = [cage_type for cage_type in ordered_cage_types(cage_values) if cage_type in cage_values]
    default_ring_sizes = result.ring_report_sizes or tuple(result.rings)
    enabled_ring_sizes = sorted(set(ring_sizes if ring_sizes is not None else default_ring_sizes))
    cpp_mode = is_cpp_mode(analysis_mode)
    metadata = input_metadata or {}
    metadata_rows = [
        [key, metadata[key]]
        for key in (
            "input_format", "topology", "trajectory_stride", "lammps_units",
            "lammps_timestep", "lammps_atom_style", "lammps_type_map_source",
        )
        if metadata.get(key) not in (None, "", "<none>")
    ]
    frame_information_rows = [
        ["sqq version", __version__],
        ["mode", mode_display(analysis_mode)],
        ["date & time", report_datetime_label()],
        ["source", source_label(result.frame.source)],
        *metadata_rows,
        ["frame", result.frame.name],
        ["time_ps", result.frame.time_ps],
        ["graph_mode", graph_mode_display(requested_bond_mode or row["connection_mode"], [row["connection_mode"]])],
        ["bond_mode", row["connection_mode"]],
        ["ring_sizes", ", ".join(str(size) for size in enabled_ring_sizes)],
    ]
    if not cpp_mode:
        frame_information_rows.append(["find_cluster", "on" if result.hydrate_cluster_enabled else "off"])
    frame_information_rows.extend([
        ["status", "ok"],
        ["n_atoms", len(result.frame.atoms)],
        ["n_waters", len(result.waters)],
        ["n_guests", len(result.guests)],
    ])
    lines = [
        f"# SQQ Frame Report: {result.frame.name}",
        "",
        *section_table("Frame Information", ["item", "value"], frame_information_rows),
    ]
    lines.extend(section_table("Molecules", ["resname", "molecules", "atoms"], molecule_count_rows(result)))
    lines.extend(connection_info_section(result, row))

    if not cpp_mode:
        ring_rows = [
            [size, row.get(f"ring{size}", 0), row.get(f"free_ring{size}", 0)]
            for size in enabled_ring_sizes
        ]
        ring_rows.append(
            [
                "total",
                sum(int(item[1]) for item in ring_rows),
                sum(int(item[2]) for item in ring_rows),
            ]
        )
        lines.extend(section_table("Ring", ["ring size", "total", "free"], ring_rows))
        lines.extend(patch_info_section("Half Cage", result.half_cages))
        lines.extend(patch_info_section("Quasi Cage", result.quasi_cages))
        lines.extend(patch_isomer_description_section("Quasi Cage Isomer Description", result.quasi_cages))

    lines.extend(cage_info_section(result, cage_types))
    lines.extend(cage_isomer_description_section(result, cage_types))
    lines.extend(cage_occupancy_section(result, cage_types, evaluated=not cpp_mode or bool(result.guests)))
    if not cpp_mode:
        lines.extend(hydrate_cluster_info_section(result))
    has_focus = result.f3f4 is not None and bool(result.f3f4.focus_resids)
    order_headers = ["metric", "count", "mean"]
    if has_focus:
        order_headers.extend(["focus_count", "focus_mean"])
    order_rows: list[list[Any]] = []
    for name, label, prefix in (
        ("f3", "F3", "F3"),
        ("f4", "F4", "F4"),
    ):
        if name not in selected_order_set:
            continue
        metric_row = [label, row.get(f"{prefix}_count"), row.get(f"{prefix}_mean")]
        if has_focus:
            metric_row.extend(
                [
                    row.get(f"{prefix}_focus_count"),
                    row.get(f"{prefix}_focus_mean"),
                ]
            )
        order_rows.append(metric_row)
    for degree in (() if cpp_mode else q_degrees_from_order_parameters(selected_order_parameters)):
        prefix = f"q{degree}"
        metric_row = [
            f"Q{degree}",
            row.get(f"{prefix}_count"),
            row.get(f"{prefix}_mean"),
        ]
        if has_focus:
            metric_row.extend(
                [
                    row.get(f"{prefix}_focus_count"),
                    row.get(f"{prefix}_focus_mean"),
                ]
            )
        order_rows.append(metric_row)
    lines.extend(section_table("Order Parameters", order_headers, order_rows))

    if not cpp_mode and result.hydrate_order is not None:
        hydrate_order_rows: list[list[Any]] = []
        if "mcg1" in selected_order_set:
            hydrate_order_rows.append(
                ["MCG-1", order_size_label(result.hydrate_order.mcg1.largest_cluster_size), result.hydrate_order.mcg1.member_type]
            )
        if "mcg3" in selected_order_set and result.hydrate_order.mcg3 is not None:
            hydrate_order_rows.append(["MCG-3", order_size_label(result.hydrate_order.mcg3.largest_cluster_size), result.hydrate_order.mcg3.member_type])
        if "dhop35" in selected_order_set:
            hydrate_order_rows.append(
                ["DHOP35", order_size_label(result.hydrate_order.dhop35.largest_cluster_size), result.hydrate_order.dhop35.member_type]
            )
        if "dhop30" in selected_order_set and result.hydrate_order.dhop30 is not None:
            hydrate_order_rows.append(["DHOP30", order_size_label(result.hydrate_order.dhop30.largest_cluster_size), result.hydrate_order.dhop30.member_type])
        lines.extend(
            section_table(
                "Hydrate Nucleation Order Parameters",
                ["parameter", "largest cluster", "member type"],
                hydrate_order_rows,
            )
        )

    if not cpp_mode:
        ice_rows = [
            ["ice_like_waters", row["ice_like_waters"]],
            ["ice_i_waters", row["ice_i_waters"]],
            ["interfacial_ice_waters", row["interfacial_ice_waters"]],
        ]
        lines.extend(section_table("Ice", ["structure", "water molecules"], ice_rows))
    if result.warnings:
        lines.extend(["", "## Warnings", ""])
        lines.extend(f"- {warning}" for warning in result.warnings)
    (frame_dir / f"{result.frame.name}_info.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def selected_order_parameters_for_result(
    result: FrameResult,
    value: Any | None,
) -> tuple[str, ...]:
    """Resolve explicit output selection or infer legacy direct-call behavior."""
    if value is not None:
        return normalize_order_parameters(value)
    inferred: list[str] = []
    if result.f3f4 is not None:
        inferred.extend(("f3", "f4"))
        inferred.extend(f"q{degree}" for degree in result.f3f4.q_degree)
    if result.hydrate_order is not None:
        inferred.extend(("mcg1", "dhop35"))
        if result.hydrate_order.mcg3 is not None:
            inferred.append("mcg3")
        if result.hydrate_order.dhop30 is not None:
            inferred.append("dhop30")
    return normalize_order_parameters(inferred or ["none"])


def order_size_label(value: int | None) -> int | str:
    """Render unavailable order parameters explicitly without conflating them with zero."""
    return "N/A" if value is None else value


def section_table(title: str, headers: list[str], rows: list[list[Any]]) -> list[str]:
    """Render one small markdown section."""
    if not rows:
        return []
    lines = ["", f"## {title}", "", markdown_rows(headers, rows).rstrip()]
    return lines


def markdown_rows(headers: list[str], rows: list[list[Any]]) -> str:
    """Render a source-aligned Markdown table using Unicode display widths."""
    text_rows = [[format_summary_cell(value) for value in row] for row in rows]
    header_text = [str(header) for header in headers]
    widths = [max(3, display_width(header)) for header in header_text]
    for row in text_rows:
        for idx, value in enumerate(row):
            widths[idx] = max(widths[idx], display_width(value))
    lines = [
        "| " + " | ".join(pad_display(header, widths[idx]) for idx, header in enumerate(header_text)) + " |",
        "| " + " | ".join("-" * widths[idx] for idx in range(len(headers))) + " |",
    ]
    for row in text_rows:
        padded = [*row, *([""] * (len(headers) - len(row)))]
        lines.append("| " + " | ".join(pad_display(padded[idx], widths[idx]) for idx in range(len(headers))) + " |")
    return "\n".join(lines) + "\n"


def display_width(value: Any) -> int:
    """Return a practical monospace width for Markdown source alignment."""
    width = 0
    for char in str(value):
        if unicodedata.combining(char):
            continue
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1
    return width


def pad_display(value: Any, width: int) -> str:
    """Right-pad text to a requested Unicode display width."""
    text = str(value)
    return text + " " * max(0, width - display_width(text))


def molecule_count_rows(result: FrameResult) -> list[list[Any]]:
    """Count molecules and atoms by residue name using source identity."""
    atom_counts = atom_resname_counts(result)
    molecule_counts: dict[str, int] = {}
    atoms = result.frame.atoms
    explicit_ids = [atom.molecule_id for atom in atoms]
    if any(value is not None for value in explicit_ids):
        if any(value is None for value in explicit_ids):
            raise ValueError("Atom records mix explicit and implicit molecule identities.")
        for resname, _ in dict.fromkeys(
            (atom.resname, int(atom.molecule_id)) for atom in atoms
        ):
            molecule_counts[resname] = molecule_counts.get(resname, 0) + 1
    else:
        previous_residue: tuple[int, str] | None = None
        for atom in atoms:
            residue = (atom.resid, atom.resname)
            if residue != previous_residue:
                molecule_counts[atom.resname] = molecule_counts.get(atom.resname, 0) + 1
                previous_residue = residue
    rows = [[resname, molecule_counts.get(resname, 0), atom_counts[resname]] for resname in atom_counts]
    rows.append(["total", sum(molecule_counts.values()), len(result.frame.atoms)])
    return rows


def connection_info_section(result: FrameResult, row: dict[str, Any]) -> list[str]:
    """Render network coordination diagnostics without modifying graph edges."""
    mode = result.graph.mode
    title, count_label = {
        "hbond": ("Hydrogen-Bond Coordination", "hydrogen bonds"),
        "oo": ("O-O Connectivity Coordination", "O-O connections"),
        "pairs": ("Pair Connectivity Coordination", "user-defined pairs"),
    }.get(mode, ("Network Coordination", "connections"))
    rows: list[list[Any]] = [
        ["water molecules", len(result.waters), ""],
        [count_label, row["connection_count"], ""],
        ["mean coordination", row["mean_coordination"], ""],
    ]
    for degree in range(5):
        rows.append([
            f"degree {degree}",
            row[f"coordination_{degree}"],
            row[f"coordination_{degree}_fraction"],
        ])
    rows.append(["degree >4", row["coordination_gt4"], row["coordination_gt4_fraction"]])
    rows.append(["degree <=2", "", row["degree_le2_fraction"]])
    return section_table(title, ["item", "count/value", "fraction"], rows)


def atom_resname_counts(result: FrameResult) -> dict[str, int]:
    """Count atoms by residue name while preserving source-file order."""
    counts: dict[str, int] = {}
    for atom in result.frame.atoms:
        counts[atom.resname] = counts.get(atom.resname, 0) + 1
    return counts


def patch_info_section(title: str, patches: list[CagePatch]) -> list[str]:
    """Render open patches as composition totals with nested isomers."""
    counts = patch_counts(patches)
    type_header = f"{title.lower().replace(' ', '-')} type"
    if not counts:
        return section_table(title, [type_header, "count"], [["total", 0]])

    grouped: dict[str, dict[str, int]] = {}
    for patch_type, count in counts.items():
        label = patch_display_label(patch_type)
        parent = label.translate(SUBSCRIPT_DIGIT_DELETE)
        children = grouped.setdefault(parent, {})
        children[label] = children.get(label, 0) + count

    rows: list[list[Any]] = []
    for parent in sorted(grouped):
        children = grouped[parent]
        rows.append([parent, sum(children.values())])
        if any(child != parent for child in children):
            labels = sorted(children)
            for index, label in enumerate(labels):
                branch = TREE_LAST if index == len(labels) - 1 else TREE_MIDDLE
                rows.append([f"{branch} {label}", f"{branch} {children[label]}"])
    rows.append(["total", sum(counts.values())])
    return section_table(title, [type_header, "count"], rows)


def patch_isomer_description_section(title: str, patches: list[CagePatch]) -> list[str]:
    """Explain each reported patch isomer as a layered ring sequence."""
    if not patches:
        return []
    counts: dict[str, int] = {}
    descriptions: dict[str, str] = {}
    for patch in patches:
        label = patch_display_label(patch.patch_type)
        counts[label] = counts.get(label, 0) + 1
        descriptions[label] = describe_patch_isomer(patch)
    rows = [[label, counts[label], descriptions[label]] for label in sorted(counts)]
    return section_table(title, ["isomer", "count", "description"], rows)


def describe_patch_isomer(patch: CagePatch) -> str:
    """Describe base/L1/L2/L3 ring layers for one quasi-cage label."""
    if not patch.layers:
        return "Layer information is not available."
    parts = [f"base ring: {patch.layers[0].removesuffix('r')}"]
    for index, layer in enumerate(patch.layers[1:], start=1):
        sequence = subscript_digit_text(layer)
        composition = layer.translate(SUBSCRIPT_DIGIT_DELETE)
        if sequence:
            layer_name = "closed side-ring sequence" if index == 1 else "outer-layer ring sequence"
            parts.append(f"L{index}: {layer_name} {sequence} ({composition})")
        else:
            parts.append(f"L{index}: composition {composition}")
    return "; ".join(parts) + "."


def subscript_digit_text(text: str) -> str:
    """Return normal digits from Unicode subscript digits in a label."""
    digits: list[str] = []
    for char in text:
        code = ord(char)
        if 0x2080 <= code <= 0x2089:
            digits.append(str(code - 0x2080))
    return "".join(digits)


def patch_display_label(patch_type: str) -> str:
    """Remove internal HC/QC prefixes from human-facing patch labels."""
    for prefix in ("hc_", "qc_"):
        if patch_type.startswith(prefix):
            return patch_type.removeprefix(prefix)
    return patch_type


def superscript_number(value: int) -> str:
    """Render small integer counts with Unicode superscript digits."""
    superscripts = str.maketrans("0123456789-", "⁰¹²³⁴⁵⁶⁷⁸⁹⁻")
    return str(value).translate(superscripts)


def ordered_cage_types(types) -> list[str]:
    """Return canonical hydrate cages first, followed by other cage labels."""
    values = set(types.keys() if isinstance(types, dict) else types)
    ordered = list(KNOWN_CAGE_TYPES)
    extras = sorted((cage_type for cage_type in values if cage_type not in KNOWN_CAGE_TYPES), key=cage_sort_key)
    return ordered + extras


def present_cage_types(types) -> list[str]:
    """Return canonical cage ordering restricted to values that are present."""
    values = set(types.keys() if isinstance(types, dict) else types)
    return [cage_type for cage_type in ordered_cage_types(values) if cage_type in values]


def cage_sort_key(cage_type: str) -> tuple[int, tuple[tuple[int, int], ...], str]:
    """Sort other cages by face count and then by label."""
    counts = parse_cage_face_label(cage_type)
    if counts is None:
        return (999, (), cage_type)
    return (sum(counts.values()), tuple(sorted(counts.items())), cage_type)


def cage_display_label(cage_type: str) -> str:
    """Render compact cage labels for human-facing markdown tables."""
    known = {
        "512": f"5{superscript_number(12)}",
        "51262": f"5{superscript_number(12)}6{superscript_number(2)}",
        "51263": f"5{superscript_number(12)}6{superscript_number(3)}",
        "51264": f"5{superscript_number(12)}6{superscript_number(4)}",
        "51268": f"5{superscript_number(12)}6{superscript_number(8)}",
        "435663": f"4{superscript_number(3)}5{superscript_number(6)}6{superscript_number(3)}",
    }
    if cage_type in known:
        return known[cage_type]
    counts = parse_cage_face_label(cage_type)
    if counts is None:
        return cage_type
    return "".join(f"{size}{superscript_number(count)}" for size, count in sorted(counts.items()) if count > 0)


def cage_occupancy_section(
    result: FrameResult,
    cage_types: list[str],
    *,
    evaluated: bool = True,
) -> list[str]:
    """Show one cage type per row with dynamic guest-composition columns."""
    if not evaluated:
        return section_table("Cage Occupancy", ["status"], [["not evaluated (no selected guests)"]])
    guests_by_id = build_guest_lookup(result.guests)
    guest_order = guest_resname_order(result)
    counts: dict[str, dict[str, int]] = {
        cage_type: {"empty": 0, "occupied": 0} for cage_type in cage_types
    }
    compositions: set[str] = set()
    for cage in result.cages:
        cage_counts = counts.setdefault(cage.cage_type, {"empty": 0, "occupied": 0})
        cage_counts["occupied" if cage.occupied else "empty"] += 1
        if not cage.occupied:
            continue
        composition = guest_composition_label(cage, guests_by_id, guest_order) or "unknown"
        compositions.add(composition)
        cage_counts[composition] = cage_counts.get(composition, 0) + 1

    guest_labels = [label for label in guest_order if label in compositions]
    extra_guest_labels = sorted(
        compositions.difference(guest_labels),
        key=lambda label: guest_composition_sort_key(label, guest_order),
    )
    child_labels = [*guest_labels, *extra_guest_labels]
    child_markers = [TREE_LAST if index == len(child_labels) - 1 else TREE_MIDDLE for index in range(len(child_labels))]
    headers = [
        "cage type",
        "total",
        "empty",
        "occupied",
        *[f"{marker} {label}" for marker, label in zip(child_markers, child_labels)],
    ]
    table_rows: list[list[Any]] = []
    for cage_type in cage_types:
        cage_counts = counts.get(cage_type, {})
        empty = cage_counts.get("empty", 0)
        occupied = cage_counts.get("occupied", 0)
        table_rows.append(
            [
                cage_display_label(cage_type),
                empty + occupied,
                empty,
                occupied,
                *[f"{marker} {cage_counts.get(label, 0)}" for marker, label in zip(child_markers, child_labels)],
            ]
        )

    total_empty = sum(counts.get(cage_type, {}).get("empty", 0) for cage_type in cage_types)
    total_occupied = sum(counts.get(cage_type, {}).get("occupied", 0) for cage_type in cage_types)
    table_rows.append(
        [
            "total",
            total_empty + total_occupied,
            total_empty,
            total_occupied,
            *[
                f"{marker} {sum(counts.get(cage_type, {}).get(label, 0) for cage_type in cage_types)}"
                for marker, label in zip(child_markers, child_labels)
            ],
        ]
    )
    return section_table("Cage Occupancy", headers, table_rows)


def guest_composition_sort_key(label: str, guest_order: list[str]) -> tuple[Any, ...]:
    """Sort exact guest compositions by occupancy size and source guest order."""
    order_index = {name: index for index, name in enumerate(guest_order)}
    components: list[tuple[int, str, int]] = []
    total_guests = 0
    for part in label.split("+"):
        name, separator, count_text = part.rpartition("x")
        if separator and count_text.isdigit():
            count = int(count_text)
        else:
            name = part
            count = 1
        total_guests += count
        components.append((order_index.get(name, 10_000), name, count))
    return total_guests, len(components), tuple(components), label


def guest_resname_order(result: FrameResult) -> list[str]:
    """Return guest residue names by their first atom position in the frame."""
    return guest_resname_order_from_guests(result.guests)


def cage_info_section(result: FrameResult, cage_types: list[str]) -> list[str]:
    """Show cage totals with nested structural isomers in one section."""
    isomers: dict[str, dict[str, int]] = {cage_type: {} for cage_type in cage_types}
    for cage in result.cages:
        label = cage.isomer or "plain"
        type_isomers = isomers.setdefault(cage.cage_type, {})
        type_isomers[label] = type_isomers.get(label, 0) + 1

    rows: list[list[Any]] = []
    for cage_type in cage_types:
        cage_label = cage_display_label(cage_type)
        type_isomers = isomers.get(cage_type, {})
        rows.append([cage_label, sum(type_isomers.values())])
        if len(type_isomers) <= 1:
            continue
        labels = sorted(type_isomers)
        for index, label in enumerate(labels):
            branch = TREE_LAST if index == len(labels) - 1 else TREE_MIDDLE
            rows.append([f"{branch} {cage_label}_{label}", f"{branch} {type_isomers[label]}"])
    rows.append(["total", len(result.cages)])
    return section_table("Cage", ["cage type", "count"], rows)


def cage_isomer_description_section(result: FrameResult, cage_types: list[str]) -> list[str]:
    """Explain each reported cage isomer as a 6-face adjacency pattern."""
    if not result.cages:
        return []
    counts: dict[tuple[str, str], int] = {}
    for cage in result.cages:
        label = cage.isomer or "plain"
        key = (cage.cage_type, label)
        counts[key] = counts.get(key, 0) + 1

    rows: list[list[Any]] = []
    for cage_type in cage_types:
        cage_label = cage_display_label(cage_type)
        labels = sorted(label for type_name, label in counts if type_name == cage_type)
        for label in labels:
            display = cage_label if label == "plain" else f"{cage_label}_{label}"
            rows.append([display, counts[(cage_type, label)], describe_cage_isomer(cage_type, label)])
    return section_table("Cage Isomer Description", ["isomer", "count", "description"], rows)


def describe_cage_isomer(cage_type: str, label: str) -> str:
    """Describe a cage isomer label in human-facing terms."""
    composition = describe_cage_face_composition(cage_type)
    arrangement = describe_hex_adjacency_label(label)
    return f"{composition}; {arrangement}"


def describe_cage_face_composition(cage_type: str) -> str:
    """Return a compact text description of cage face counts."""
    counts = parse_cage_face_label(cage_type)
    if not counts:
        return f"face composition: {cage_display_label(cage_type)}"
    parts = [
        f"{count} {size}-ring face{'s' if count != 1 else ''}"
        for size, count in sorted(counts.items())
        if count > 0
    ]
    return "face composition: " + ", ".join(parts)


def describe_hex_adjacency_label(label: str) -> str:
    """Return a readable explanation of the 6-ring face adjacency label."""
    if label == "plain":
        return "no 6-ring face arrangement isomer is reported"
    descriptions = {
        "6single": "one 6-ring face; no 6-6 shared edge is possible",
        "6adj": "two 6-ring faces share one edge",
        "6pair+single": "three 6-ring faces contain one adjacent pair and one separated single face",
        "6chain3": "three 6-ring faces form a chain with two 6-6 shared edges",
        "6tri3": "three 6-ring faces are mutually adjacent",
        "6pair+2single": "four 6-ring faces contain one adjacent pair and two separated single faces",
        "2x6pair": "four 6-ring faces form two separated adjacent pairs",
        "6chain3+single": "four 6-ring faces contain one three-face chain and one separated single face",
        "6star3": "four 6-ring faces form a star: one face touches three others",
        "6chain4": "four 6-ring faces form a four-face chain",
        "6tri3+single": "four 6-ring faces contain one mutually adjacent triple and one separated single face",
        "6cycle4": "four 6-ring faces form a four-face cycle",
        "6tri3+tail": "four 6-ring faces contain one mutually adjacent triple with one attached tail face",
        "6K4-e": "four 6-ring faces are almost fully connected, with one 6-6 adjacency missing",
        "6K4": "four 6-ring faces are all mutually adjacent",
    }
    if label in descriptions:
        return descriptions[label]
    separated = re.fullmatch(r"(\d+)x6sep", label)
    if separated:
        return f"{separated.group(1)} 6-ring faces are all separated from each other"
    generic = re.fullmatch(r"6n(\d+)e(\d+)d(\d+)", label)
    if generic:
        n_hex, edge_count, degree_text = generic.groups()
        return f"{n_hex} 6-ring faces with {edge_count} shared 6-6 edges; degree sequence {degree_text}"
    return f"6-ring face arrangement label: {label}"


def source_label(source: Path | None) -> str:
    """Return an absolute source path for human-facing reports."""
    if source is None:
        return ""
    return str(Path(source).resolve())


def report_datetime_label(value: datetime | None = None) -> str:
    """Return local report-generation time for per-frame Markdown output."""
    moment = value or datetime.now().astimezone()
    zone = report_timezone_label(moment)
    return f"{moment:%Y-%m-%d %H:%M:%S} {zone}"


def report_timezone_label(value: datetime) -> str:
    """Return a compact human-facing timezone label."""
    name = value.tzname() or "UTC"
    offset = value.utcoffset()
    total_minutes = int(offset.total_seconds() / 60) if offset is not None else 0
    if total_minutes == 480 and name in {"CST", "China Standard Time", "\u4e2d\u56fd\u6807\u51c6\u65f6\u95f4"}:
        return "Asia/Shanghai"
    if offset is None:
        return name
    sign = "+" if total_minutes >= 0 else "-"
    hours, minutes = divmod(abs(total_minutes), 60)
    offset_text = f"UTC{sign}{hours:02d}:{minutes:02d}"
    return name if name and name != "UTC" else offset_text


def write_membership(result: FrameResult, frame_dir: Path) -> None:
    """Write object-to-water membership for plotting and debugging."""
    rows: list[dict[str, Any]] = []
    water_resid_by_oxygen = {water.oxygen: water.resid for water in result.waters}
    reported_ring_sizes = set(result.ring_report_sizes or tuple(result.rings))
    for size, rings in sorted(result.rings.items()):
        if size not in reported_ring_sizes:
            continue
        for ring in rings:
            rows.append(
                {
                    "object_id": ring.object_id,
                    "object_type": f"ring{size}",
                    "center_atom_name": f"R{size}",
                    "oxygen_indices": ",".join(str(idx) for idx in ring.nodes),
                    "water_resids": ",".join(str(water_resid_by_oxygen[idx]) for idx in ring.nodes),
                    "guest_ids": "",
                    "isomer": "",
                    "unwrap_conflict": "false",
                }
            )
    for patch in [*result.half_cages, *result.quasi_cages]:
        rows.append(
            {
                "object_id": patch.object_id,
                "object_type": patch.patch_type,
                "center_atom_name": "HC" if patch.kind == "half_cage" else "QC",
                "oxygen_indices": ",".join(str(idx) for idx in patch.waters),
                "water_resids": ",".join(str(water_resid_by_oxygen[idx]) for idx in patch.waters),
                "guest_ids": "",
                "isomer": "",
                "unwrap_conflict": "false",
            }
        )
    for cage in result.cages:
        rows.append(
            {
                "object_id": cage.object_id,
                "object_type": cage.cage_type,
                "center_atom_name": cage_center_name(cage.cage_type),
                "oxygen_indices": ",".join(str(idx) for idx in cage.waters),
                "water_resids": ",".join(str(water_resid_by_oxygen[idx]) for idx in cage.waters),
                "guest_ids": ",".join(cage.guest_ids),
                "isomer": cage.isomer or "",
                "unwrap_conflict": "false",
            }
        )
    data = pd.DataFrame(rows)
    data.to_csv(frame_dir / f"{result.frame.name}_membership.tsv", sep="\t", index=False)


def write_order_parameter(
    result: FrameResult,
    frame_dir: Path,
    order_parameters: Any | None = None,
) -> None:
    """Write per-water F3/F4/Q_l values for custom plotting or focus-water checks."""
    path = frame_dir / f"{result.frame.name}_order_parameter.tsv"
    selected = selected_order_parameters_for_result(result, order_parameters)
    selected_set = set(selected)
    q_degrees = q_degrees_from_order_parameters(selected)
    if result.f3f4 is None or not (selected_set & {"f3", "f4"} or q_degrees):
        path.unlink(missing_ok=True)
        return
    rows = []
    for item in result.f3f4.per_water:
        row = {
            "resid": item.resid,
            "atomid": item.atomid,
            "oxygen_index": item.oxygen,
            "x_nm": item.xyz[0],
            "y_nm": item.xyz[1],
            "z_nm": item.xyz[2],
        }
        if "f3" in selected_set:
            row["F3"] = item.f3
        if "f4" in selected_set:
            row["F4"] = item.f4
        for degree in q_degrees:
            row[f"q{degree}"] = item.q_values.get(degree)
        if q_degrees:
            row["q_neighbors"] = item.q_neighbors
        rows.append(row)
    pd.DataFrame(rows).to_csv(path, sep="\t", index=False)


def cage_center_name(cage_type: str) -> str:
    """Return the short CNT atom name used for a cage center."""
    return {"512": "G512", "51262": "G62", "51263": "G63", "51264": "G64", "51268": "G68", "435663": "G436"}.get(cage_type, "CAGE")[:5]


def write_vmd_script(result: FrameResult, frame_dir: Path) -> None:
    """Write a small VMD helper script with default colors."""
    path = frame_dir / f"{result.frame.name}_view.vmd.tcl"
    lines = [
        "# SQQ VMD helper",
        "# Source this after loading SQQ GRO files.",
        "color Name R4 gray",
        "color Name R5 purple",
        "color Name R6 tan",
        "color Name R7 black",
        "color Name CP5 cyan",
        "color Name CP6 yellow",
        "color Name G512 blue",
        "color Name G62 green",
        "color Name G63 orange",
        "color Name G64 red",
        "display projection Orthographic",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(
    rows: list[dict[str, Any]],
    outdir: Path,
    config: dict[str, Any],
    write_xlsx: bool = True,
    run_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Write global summaries and return deterministic output timing metadata."""
    started = perf_counter()
    outdir.mkdir(parents=True, exist_ok=True)
    metrics: dict[str, Any] = {
        "dataframe_seconds": 0.0,
        "run_config_initial_seconds": 0.0,
        "detail_table_build_seconds": 0.0,
        "detail_csv": {"enabled": False, "total_seconds": 0.0, "tables": []},
        "summary_csv": {"enabled": False, "total_seconds": 0.0, "tables": []},
        "xlsx": {"enabled": False, "total_seconds": 0.0, "sheets": []},
    }

    dataframe_started = perf_counter()
    columns = list(SUMMARY_COLUMNS)
    extra_columns = stable_extra_columns(rows, columns)
    public_rows = [
        {key: value for key, value in row.items() if not str(key).startswith("_")}
        for row in rows
    ]
    data = pd.DataFrame(public_rows, columns=columns + extra_columns)
    metrics["dataframe_seconds"] = round(perf_counter() - dataframe_started, 6)

    summary_md = outdir / "summary.md"
    if summary_md.exists():
        summary_md.unlink()

    config_started = perf_counter()
    config_for_dump = write_run_config(outdir, config, run_info or {})
    metrics["run_config_initial_seconds"] = round(perf_counter() - config_started, 6)

    detail_index = pd.DataFrame()
    if (
        output_enabled(config, "summary-detail-csv")
        or output_enabled(config, "cluster-detail")
    ):
        detail_build_started = perf_counter()
        detail_tables = summary_detail_tables(data, config, raw_rows=rows)
        metrics["detail_table_build_seconds"] = round(perf_counter() - detail_build_started, 6)
        detail_index, detail_metrics = write_summary_detail_csvs(
            outdir,
            detail_tables,
            config,
            return_metrics=True,
        )
        metrics["detail_csv"] = detail_metrics
    else:
        remove_summary_detail_csvs(outdir, config)

    write_summary_csv = output_enabled(config, "summary-csv")
    write_summary_xlsx = bool(write_xlsx) and output_enabled(config, "summary-xlsx")
    summary_tables: list[tuple[str, pd.DataFrame, bool]] = []
    if write_summary_csv or write_summary_xlsx:
        summary_tables = summary_output_tables(
            data,
            config,
            run_info or {},
            config_for_dump,
            detail_index,
        )

    if write_summary_csv:
        metrics["summary_csv"] = write_summary_csvs(
            outdir, summary_tables, config, return_metrics=True
        )
    else:
        remove_summary_csvs(outdir, config)

    summary_xlsx = outdir / "summary.xlsx"
    if write_summary_xlsx:
        xlsx_started = perf_counter()
        xlsx_metrics: dict[str, Any] = {
            "enabled": True,
            "table_write_seconds": 0.0,
            "format_seconds": 0.0,
            "save_seconds": 0.0,
            "total_seconds": 0.0,
            "bytes": 0,
            "sheets": [],
        }
        temp_path = temporary_output_path(summary_xlsx)
        writer: pd.ExcelWriter | None = None
        try:
            writer = pd.ExcelWriter(temp_path, engine="openpyxl")
            tables = summary_tables

            for sheet_name, table, include_header in tables:
                ensure_excel_table_size(table, sheet_name, include_header=include_header)
                table_started = perf_counter()
                table.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=include_header,
                )
                elapsed = perf_counter() - table_started
                xlsx_metrics["table_write_seconds"] += elapsed
                xlsx_metrics["sheets"].append(
                    table_metric(
                        sheet_name,
                        table,
                        write_seconds=elapsed,
                    )
                )

            format_started = perf_counter()
            format_metrics = format_summary_workbook(writer.book)
            xlsx_metrics["format_seconds"] = perf_counter() - format_started
            format_by_sheet = {item["sheet"]: item for item in format_metrics}
            for item in xlsx_metrics["sheets"]:
                item.update(format_by_sheet.get(item["sheet"], {}))

            save_started = perf_counter()
            writer.close()
            writer = None
            xlsx_metrics["save_seconds"] = perf_counter() - save_started
            os.replace(temp_path, summary_xlsx)
            xlsx_metrics["bytes"] = summary_xlsx.stat().st_size
        except Exception:
            if writer is not None:
                try:
                    writer.close()
                except Exception:
                    pass
            raise
        finally:
            temp_path.unlink(missing_ok=True)
        xlsx_metrics["table_write_seconds"] = round(xlsx_metrics["table_write_seconds"], 6)
        xlsx_metrics["format_seconds"] = round(xlsx_metrics["format_seconds"], 6)
        xlsx_metrics["save_seconds"] = round(xlsx_metrics["save_seconds"], 6)
        xlsx_metrics["total_seconds"] = round(perf_counter() - xlsx_started, 6)
        metrics["xlsx"] = xlsx_metrics
    else:
        summary_xlsx.unlink(missing_ok=True)

    metrics["total_seconds"] = round(perf_counter() - started, 6)
    return metrics


def summary_output_tables(
    data: pd.DataFrame,
    config: dict[str, Any],
    run_info: dict[str, Any],
    config_for_dump: dict[str, Any],
    detail_index: pd.DataFrame,
) -> list[tuple[str, pd.DataFrame, bool]]:
    """Build the shared table list for summary XLSX and CSV outputs."""
    tables: list[tuple[str, pd.DataFrame, bool]] = [
        ("summary", summary_dashboard_table(data, run_info, config), False),
    ]
    tables.extend(
        (sheet_name, table, True)
        for sheet_name, table in summary_sheet_tables(data, config).items()
    )
    if not is_cpp_mode(config.get("mode", "50")) and not detail_index.empty:
        tables.append(("detail_index", detail_index, True))
    tables.append(("config", pd.DataFrame(flatten_config(config_for_dump)), True))
    return tables


def write_summary_csvs(
    outdir: Path,
    tables: list[tuple[str, pd.DataFrame, bool]],
    config: dict[str, Any],
    *,
    return_metrics: bool = False,
) -> dict[str, Any] | None:
    """Atomically write one main-summary CSV per workbook-equivalent table."""
    started = perf_counter()
    dir_name = (
        str(config.get("output", {}).get("summary_csv_dir", "summary")).strip()
        or "summary"
    )
    summary_dir = outdir / dir_name
    summary_dir.mkdir(parents=True, exist_ok=True)
    metrics: dict[str, Any] = {
        "enabled": True,
        "total_seconds": 0.0,
        "tables": [],
    }
    pending: list[tuple[Path, Path]] = []
    written_names: set[str] = set()
    try:
        for name, table, include_header in tables:
            target = summary_dir / f"{name}.csv"
            temp_path = temporary_output_path(target)
            pending.append((temp_path, target))
            write_started = perf_counter()
            table.to_csv(
                temp_path,
                index=False,
                header=include_header,
                encoding="utf-8-sig",
            )
            elapsed = perf_counter() - write_started
            written_names.add(name)
            metric = table_metric(name, table, write_seconds=elapsed)
            metric["bytes"] = temp_path.stat().st_size
            metrics["tables"].append(metric)

        stale_paths = [
            summary_dir / f"{name}.csv"
            for name in SUMMARY_MAIN_TABLE_NAMES
            if name not in written_names
        ]
        commit_output_bundle(pending, stale_paths)
    finally:
        for temp_path, _ in pending:
            temp_path.unlink(missing_ok=True)

    metrics["total_seconds"] = round(perf_counter() - started, 6)
    if return_metrics:
        return metrics
    return None


def remove_summary_csvs(outdir: Path, config: dict[str, Any]) -> None:
    """Remove known main-summary CSVs while preserving unrelated files."""
    dir_name = (
        str(config.get("output", {}).get("summary_csv_dir", "summary")).strip()
        or "summary"
    )
    summary_dir = outdir / dir_name
    if not summary_dir.exists():
        return
    commit_output_bundle(
        [],
        [summary_dir / f"{name}.csv" for name in SUMMARY_MAIN_TABLE_NAMES],
    )
    try:
        summary_dir.rmdir()
    except OSError:
        pass


def temporary_output_path(target: Path) -> Path:
    """Create a same-directory temporary path suitable for atomic replacement."""
    target.parent.mkdir(parents=True, exist_ok=True)
    descriptor, raw_path = tempfile.mkstemp(
        prefix=f".{target.stem}.",
        suffix=target.suffix,
        dir=target.parent,
    )
    os.close(descriptor)
    return Path(raw_path)


def ensure_excel_table_size(
    table: pd.DataFrame,
    sheet_name: str,
    *,
    include_header: bool = True,
) -> None:
    """Fail early with a SQQ-specific message before pandas reaches Excel limits."""
    rows = len(table) + int(include_header)
    columns = len(table.columns)
    if rows > EXCEL_MAX_ROWS or columns > EXCEL_MAX_COLUMNS:
        raise ValueError(
            f"Summary sheet {sheet_name!r} is too large for Excel "
            f"({rows} rows x {columns} columns; limits are "
            f"{EXCEL_MAX_ROWS} rows x {EXCEL_MAX_COLUMNS} columns). "
            "Use summary-detail-csv for high-cardinality detail data."
        )


def commit_output_bundle(
    pending: list[tuple[Path, Path]],
    removals: list[Path] | tuple[Path, ...] = (),
) -> None:
    """Replace a related output group together, restoring prior files on failure."""
    targets = list(dict.fromkeys([target for _, target in pending] + list(removals)))
    backups: dict[Path, Path] = {}
    committed: list[Path] = []
    try:
        for target in targets:
            if not target.exists():
                continue
            backup = temporary_output_path(target)
            backup.unlink(missing_ok=True)
            os.replace(target, backup)
            backups[target] = backup
        for temp_path, target in pending:
            os.replace(temp_path, target)
            committed.append(target)
    except Exception:
        for target in committed:
            target.unlink(missing_ok=True)
        for target, backup in backups.items():
            if backup.exists():
                os.replace(backup, target)
        raise
    finally:
        for backup in backups.values():
            backup.unlink(missing_ok=True)


def table_metric(sheet_name: str, table: pd.DataFrame, *, write_seconds: float = 0.0) -> dict[str, Any]:
    """Return compact, YAML-safe size and write metadata for one output table."""
    rows, columns = table.shape
    return {
        "sheet": sheet_name,
        "rows": int(rows),
        "columns": int(columns),
        "cells": int(rows * columns),
        "write_seconds": round(float(write_seconds), 6),
    }


def write_run_config(
    outdir: Path,
    config: dict[str, Any],
    run_info: dict[str, Any],
) -> dict[str, Any]:
    """Atomically write mandatory run metadata and return the dumped mapping."""
    outdir.mkdir(parents=True, exist_ok=True)
    config_for_dump = config_with_run_metadata(config, run_info)
    target = outdir / "run_config.yaml"
    temp_path = temporary_output_path(target)
    try:
        with temp_path.open("w", encoding="utf-8", newline="\n") as handle:
            dump_config(config_for_dump, handle)
        os.replace(temp_path, target)
    finally:
        temp_path.unlink(missing_ok=True)
    return config_for_dump


def config_with_run_metadata(config: dict[str, Any], run_info: dict[str, Any]) -> dict[str, Any]:
    """Return a config copy that preserves raw settings and records resolved run metadata."""
    if not run_info:
        return config
    enriched = deepcopy(config)
    run = enriched.setdefault("run", {})
    parallel = config.get("parallel", {})
    run.update({
        "sqq_version": run_info.get("sqq_version", __version__),
        "status": run_info.get("status", ""),
        "error": run_info.get("error", ""),
        "graph_mode_requested": run_info.get("graph_mode", config.get("graph", {}).get("bond_mode", "")),
        "graph_mode_effective": run_info.get("effective_graph_modes", ""),
        "graph_mode_display": run_info.get("graph_mode_display", ""),
        "order_parameters": run_info.get(
            "order_parameters",
            order_parameter_display(config.get("order", {}).get("parameters")),
        ),
        "find_cluster": run_info.get(
            "find_cluster",
            "on" if config.get("hydrate_cluster", {}).get("enabled", False) else "off",
        ),
        "output_types": run_info.get(
            "output_types",
            output_type_display(
                config.get("output", {}).get("types"),
                cpp_mode=is_cpp_mode(config.get("mode", "50")),
            ),
        ),
        "frames_total": run_info.get("frames_total", ""),
        "frames_ok": run_info.get("frames_ok", ""),
        "frames_failed": run_info.get("frames_failed", ""),
        "failures": run_info.get("failures", []),
        "worker_request": parallel.get("workers", "auto"),
        "worker_policy": run_info.get("worker_policy", ""),
        "workers_resolved": run_info.get("workers", ""),
        "parallel_backend": run_info.get("parallel_backend", "serial"),
        "math_threads_per_worker": run_info.get("math_threads", 1),
        "summary_write": run_info.get("summary_write", {}),
    })
    for key in (
        "topology_group_count",
        "topology_group_limit",
        "topology_group_limit_exceeded",
        "topology_group_labels_enabled",
        "info_only_fallback_required",
        "topology_grouping",
        "topology_groups",
        "topology_source_mapping",
        "topology_group",
        "topology_fingerprint",
        "requested_output_types",
        "output_policy",
        "warnings",
    ):
        if key in run_info:
            run[key] = deepcopy(run_info[key])
    return enriched


def write_summary_detail_csvs(
    outdir: Path,
    tables: dict[str, pd.DataFrame],
    config: dict[str, Any],
    *,
    return_metrics: bool = False,
) -> pd.DataFrame | tuple[pd.DataFrame, dict[str, Any]]:
    """Atomically write detail CSVs and optionally return their timing metadata."""
    started = perf_counter()
    detail_dir_name = str(config.get("output", {}).get("summary_detail_dir", "summary_detail")).strip() or "summary_detail"
    detail_dir = outdir / detail_dir_name
    detail_dir.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, Any]] = []
    metrics: dict[str, Any] = {"enabled": True, "total_seconds": 0.0, "tables": []}
    pending: list[tuple[Path, Path]] = []
    written_names: set[str] = set()
    try:
        for name in SUMMARY_DETAIL_TABLE_NAMES:
            table = tables.get(name)
            if table is None:
                continue
            path = detail_dir / f"{name}.csv"
            temp_path = temporary_output_path(path)
            pending.append((temp_path, path))
            write_started = perf_counter()
            table.to_csv(temp_path, index=False, encoding="utf-8-sig")
            elapsed = perf_counter() - write_started
            written_names.add(name)
            relative_dir = detail_dir_name.rstrip("/\\")
            relative_file = f"{relative_dir}/{name}.csv".replace("\\", "/")
            rows.append(
                {
                    "table": name,
                    "file": relative_file,
                    "rows": len(table),
                    "columns": len(table.columns),
                }
            )
            metric = table_metric(name, table, write_seconds=elapsed)
            metric["bytes"] = temp_path.stat().st_size
            metrics["tables"].append(metric)
        stale_paths = [
            detail_dir / f"{name}.csv"
            for name in SUMMARY_DETAIL_TABLE_NAMES
            if name not in written_names
        ]
        commit_output_bundle(pending, stale_paths)
    finally:
        for temp_path, _ in pending:
            temp_path.unlink(missing_ok=True)
    metrics["total_seconds"] = round(perf_counter() - started, 6)
    detail_index = pd.DataFrame(rows, columns=["table", "file", "rows", "columns"])
    if return_metrics:
        return detail_index, metrics
    return detail_index


def remove_summary_detail_csvs(outdir: Path, config: dict[str, Any]) -> None:
    """Remove known stale detail CSVs when summary-detail-csv is disabled."""
    detail_dir_name = str(
        config.get("output", {}).get("summary_detail_dir", "summary_detail")
    ).strip() or "summary_detail"
    detail_dir = outdir / detail_dir_name
    if not detail_dir.exists():
        return
    commit_output_bundle(
        [],
        [detail_dir / f"{name}.csv" for name in SUMMARY_DETAIL_TABLE_NAMES],
    )
    try:
        detail_dir.rmdir()
    except OSError:
        pass


def summary_dashboard_table(data: pd.DataFrame, run_info: dict[str, Any], config: dict[str, Any]) -> pd.DataFrame:
    """Build a compact human-facing dashboard for the summary sheet."""
    cpp_mode = is_cpp_mode(config.get("mode", "50"))
    banner_lines = [line.strip("| ") for line in SQQ_BANNER.splitlines() if line.startswith("|")]
    title = banner_lines[0] if banner_lines else "Shell  Quant  Qualifier"
    author = banner_lines[1] if len(banner_lines) > 1 else "by J. PANG & Q. SUN"
    matched_files = run_info.get("matched_files", "")
    try:
        matched_count = int(matched_files)
    except (TypeError, ValueError):
        matched_count = 0
    graph_mode_value = run_info.get("graph_mode_display") or graph_mode_display(
        run_info.get("graph_mode", config.get("graph", {}).get("bond_mode", "auto")),
        data.get("connection_mode", []),
    )
    selected_order_parameters = normalize_order_parameters(
        config.get("order", {}).get("parameters", ["f3", "f4"])
    )
    selected_order_set = set(selected_order_parameters)
    q_degrees = q_degrees_from_order_parameters(selected_order_parameters)

    rows: list[list[Any]] = [
        [title, ""],
        [author, ""],
        ["", ""],
        ["Basic Information", ""],
        ["Date", run_info.get("date", "")],
        ["Start time", run_info.get("start_time", "")],
        ["Finish time", run_info.get("finish_time", "")],
        ["Time zone", run_info.get("time_zone", "")],
        ["Duration (s)", run_info.get("elapsed_seconds", "")],
        ["Working directory", run_info.get("working_dir", "")],
        ["Input", run_info.get("input", "")],
        ["Matched files", matched_files],
        ["Input format", run_info.get("input_format", "")],
    ]
    if matched_count > 1:
        rows.extend([
            ["First file", run_info.get("first_file", "")],
            ["Last file", run_info.get("last_file", "")],
        ])
    else:
        rows.append(["Current file", run_info.get("first_file", "")])
    rows.extend([
        ["Output directory", run_info.get("output_dir", "")],
        ["summary.xlsx", run_info.get("summary_xlsx", "")],
        ["summary_csv", run_info.get("summary_csv", "")],
        ["summary_detail_csv", run_info.get("summary_detail_csv", "")],
        ["run_config.yaml", run_info.get("run_config", "")],
        ["", ""],
        ["Configuration", ""],
        ["SQQ version", run_info.get("sqq_version", __version__)],
        ["Mode", mode_display(config.get("mode", "50"))],
        ["Config file", run_info.get("config_file", "<built-in defaults>")],
        ["Topology", run_info.get("topology", "<none>")],
        ["Graph mode", graph_mode_value],
        ["Search sizes", excel_scalar(config.get("ring", {}).get("sizes", ""))],
    ])
    if not cpp_mode:
        rows.append(["Ring report sizes", excel_scalar(configured_ring_report_sizes(config))])
    rows.append(["Ring definition", config.get("ring", {}).get("definition", "chordless")])
    rows.append(["Trajectory stride", run_info.get("trajectory_stride", 1)])
    if str(run_info.get("input_format", "")).startswith("lammps-"):
        rows.extend([
            ["LAMMPS units", run_info.get("lammps_units", "")],
            ["LAMMPS timestep", run_info.get("lammps_timestep", "")],
            ["LAMMPS atom style", run_info.get("lammps_atom_style", "")],
            ["LAMMPS type map", run_info.get("lammps_type_map_source", "")],
        ])
    if not cpp_mode:
        rows.extend([
            ["Quasi-cage sizes", f"{excel_scalar(config.get('quasi_cage', {}).get('base_sizes', 'auto'))} / {excel_scalar(config.get('quasi_cage', {}).get('side_sizes', 'auto'))}"],
            ["Quasi max layer", config.get("quasi_cage", {}).get("max_layers", "")],
            ["Quasi search policy", config.get("quasi_cage", {}).get("search_policy", "bounded")],
        ])
    rows.extend([
        ["Cage report types", dashboard_cage_targets(config)],
        ["Maximum cage face", config.get("cage", {}).get("max_faces", 20)],
    ])
    if not cpp_mode:
        rows.extend([
            ["Find cluster", "on" if config.get("hydrate_cluster", {}).get("enabled", False) else "off"],
            ["Cluster min cage", config.get("hydrate_cluster", {}).get("min_cage", 2)],
        ])
    rows.append(["Order parameters", order_parameter_display(selected_order_parameters)])
    if not cpp_mode and selected_order_set & {"mcg1", "mcg3"}:
        rows.append([
            "MCG guest / water cutoff (nm)",
            f"{config.get('hydrate_order', {}).get('mcg_guest_cutoff_nm', 0.90)} / "
            f"{config.get('hydrate_order', {}).get('mcg_water_cutoff_nm', 0.60)}",
        ])
    if not cpp_mode and selected_order_set & {"dhop35", "dhop30"}:
        rows.append([
            "DHOP O-O cutoff (nm)",
            config.get("hydrate_order", {}).get("dhop_neighbor_cutoff_nm", 0.35),
        ])
    if not cpp_mode and q_degrees:
        rows.extend([
            ["Q_l degree", excel_scalar(q_degrees)],
            ["Q_l neighbor mode", config.get("order", {}).get("q_neighbor_mode", "graph")],
            ["Q_l cutoff (nm)", config.get("order", {}).get("q_cutoff_nm", 0.35)],
            ["Q_l n neighbor", config.get("order", {}).get("q_n_neighbor", "NULL")],
        ])
    rows.extend([
        [
            "Output types",
            run_info.get("output_types")
            or output_type_display(
                config.get("output", {}).get("types"),
                cpp_mode=cpp_mode,
            ),
        ],
        ["Output layout", run_info.get("output_layout", "")],
        ["Worker policy", run_info.get("worker_policy", "")],
        ["Parallel backend", run_info.get("parallel_backend", "serial")],
        ["Math threads per worker", run_info.get("math_threads", 1)],
        ["Workers", run_info.get("workers", "")],
        ["", ""],
        ["Analysis Results (min / mean / max)", ""],
        ["Frames total / ok / failed", f"{len(data)} / {frames_ok_count(data)} / {frames_failed_count(data)}"],
        ["Water molecules", min_mean_max_column(data, "n_waters")],
        ["Guest molecules", min_mean_max_column(data, "n_guests")],
        ["Connections", min_mean_max_column(data, "connection_count")],
    ])
    if not cpp_mode:
        for size in configured_ring_report_sizes(config):
            rows.append([f"Ring{size}", min_mean_max_column(data, f"ring{size}")])
        rows.extend([
            ["Half cage", min_mean_max_column(data, "half_cage_total")],
            ["Quasi cage", min_mean_max_column(data, "quasi_cage_total")],
        ])
    rows.append(["Cage total", min_mean_max_column(data, "cage_total")])
    if cpp_mode and not has_selected_guests(data):
        rows.append(["Cage occupancy", "not evaluated"])
    else:
        rows.extend([
            ["Empty cage", min_mean_max_column(data, "cage_empty")],
            ["Occupied cage", min_mean_max_column(data, "cage_occupied")],
        ])
    if not cpp_mode and config.get("hydrate_cluster", {}).get("enabled", False):
        rows.extend([
            ["Hydrate cluster", min_mean_max_column(data, "hydrate_cluster_count")],
            ["Isolated cage", min_mean_max_column(data, "isolated_cage_count")],
        ])
    if not cpp_mode:
        rows.append(["Ice-like waters", min_mean_max_column(data, "ice_like_waters")])
    return pd.DataFrame(rows)


def frames_ok_count(data: pd.DataFrame) -> int:
    """Count successfully analyzed frames."""
    return int((data.get("status") == "ok").sum()) if "status" in data else len(data)


def frames_failed_count(data: pd.DataFrame) -> int:
    """Count failed frames."""
    return int((data.get("status") == "failed").sum()) if "status" in data else 0


def has_selected_guests(data: pd.DataFrame) -> bool:
    """Return whether any analyzed frame contains a selected guest molecule."""
    if "n_guests" not in data.columns:
        return False
    return bool((pd.to_numeric(data["n_guests"], errors="coerce").fillna(0) > 0).any())


def first_data_value(data: pd.DataFrame, column: str, fallback: Any = "") -> Any:
    """Return the first non-empty value in a summary column."""
    if column not in data:
        return fallback
    for value in data[column]:
        if pd.notna(value) and value != "":
            return value
    return fallback


def sum_numeric_column(data: pd.DataFrame, column: str) -> int:
    """Sum a numeric summary column while ignoring blanks."""
    if column not in data:
        return 0
    return int(pd.to_numeric(data[column], errors="coerce").fillna(0).sum())


def min_mean_max_column(data: pd.DataFrame, column: str) -> str:
    """Render per-frame min / mean / max statistics for a numeric summary column."""
    if column not in data:
        return "0 / 0.0 / 0"
    values = pd.to_numeric(data[column], errors="coerce").dropna()
    if values.empty:
        return "0 / 0.0 / 0"
    return " / ".join([
        format_stat_value(values.min()),
        format_stat_value(values.mean(), force_decimal=True),
        format_stat_value(values.max()),
    ])


def format_stat_value(value: Any, *, force_decimal: bool = False) -> str:
    """Format dashboard min/mean/max values compactly but readably."""
    numeric = float(value)
    if force_decimal and numeric.is_integer():
        return f"{numeric:.1f}"
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.3f}".rstrip("0").rstrip(".")


def molecule_totals(data: pd.DataFrame) -> dict[str, int]:
    """Collect total atom counts by residue name for the dashboard."""
    totals: dict[str, int] = {}
    for column in data.columns:
        if not column.startswith("mol_") or column == "mol_TOTAL":
            continue
        totals[column.removeprefix("mol_")] = sum_numeric_column(data, column)
    if "mol_TOTAL" in data:
        totals["TOTAL"] = sum_numeric_column(data, "mol_TOTAL")
    return totals


def configured_ring_report_sizes(config: dict[str, Any]) -> list[int]:
    """Return normalized ring report sizes for dashboards and data sheets."""
    search_sizes = config.get("ring", {}).get("sizes", [5, 6])
    value = config.get("ring", {}).get("report_sizes", "auto")
    if value in (None, "", "auto"):
        value = search_sizes
    if isinstance(value, str):
        return sorted({int(item.strip()) for item in value.split(",") if item.strip()})
    return sorted({int(item) for item in value})


def dashboard_cage_targets(config: dict[str, Any]) -> str:
    """Render exact cage report types with human-facing superscripts."""
    targets = config.get("cage", {}).get("report_types", [])
    if isinstance(targets, str):
        if targets.strip().lower() in {"auto", "all"}:
            return "all detected cages (follows --size)"
        raw_targets = [item.strip() for item in targets.split(",") if item.strip()]
    else:
        raw_targets = [str(item) for item in targets or []]
    return ", ".join(cage_display_label(target) for target in raw_targets)

def format_summary_workbook(workbook) -> list[dict[str, Any]]:
    """Apply formatting and return per-sheet timing and size information."""
    metrics: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        started = perf_counter()
        worksheet.sheet_view.showGridLines = False
        if worksheet.title == "summary":
            format_summary_dashboard_sheet(worksheet)
            format_mode = "dashboard"
        else:
            format_mode = format_table_sheet(worksheet)
        metrics.append(
            {
                "sheet": worksheet.title,
                "rows": int(worksheet.max_row),
                "columns": int(worksheet.max_column),
                "cells": int(worksheet.max_row * worksheet.max_column),
                "format_mode": format_mode,
                "format_seconds": round(perf_counter() - started, 6),
            }
        )
    return metrics


def format_summary_dashboard_sheet(worksheet) -> None:
    """Style the human-facing dashboard sheet."""
    title_fill = PatternFill("solid", fgColor="F7E7C6")
    author_fill = PatternFill("solid", fgColor="FFF7E6")
    section_fill = PatternFill("solid", fgColor="2563EB")
    label_fill = PatternFill("solid", fgColor="EFF6FF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    widths = {"A": 28, "B": 120}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 22
    worksheet.freeze_panes = "A4"

    worksheet.merge_cells("A1:B1")
    worksheet.merge_cells("A2:B2")
    for row in (1, 2):
        for cell in worksheet[row]:
            cell.fill = title_fill if row == 1 else author_fill
            cell.font = Font(color="3B2A14", bold=True, size=18 if row == 1 else 11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    section_labels = {"Basic Information", "Configuration", "Analysis Results (min / mean / max)"}
    for row in worksheet.iter_rows():
        if row[0].row <= 2:
            continue
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if cell.value not in (None, ""):
                cell.border = thin_border
        if row[0].value in section_labels:
            for cell in row:
                cell.fill = section_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            if row[0].value not in (None, ""):
                row[0].fill = label_fill
                row[0].font = Font(bold=True, color="0F172A")
            if row[1].value not in (None, ""):
                row[1].font = Font(color="111827")
                row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def format_table_header(worksheet) -> None:
    """Apply the shared data-sheet header style."""
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def format_table_sheet(worksheet) -> str:
    """Style a data sheet, using lightweight formatting for very large tables."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return "empty"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    format_table_header(worksheet)
    cells = worksheet.max_row * worksheet.max_column
    if cells > FULL_TABLE_FORMAT_MAX_CELLS or worksheet.max_column > FULL_TABLE_FORMAT_MAX_COLUMNS:
        for column_index, cell in enumerate(worksheet[1], start=1):
            letter = get_column_letter(column_index)
            header_width = len(str(cell.value or "")) + 2
            worksheet.column_dimensions[letter].width = min(
                max(10, header_width),
                LIGHTWEIGHT_TABLE_COLUMN_WIDTH,
            )
        return "lightweight"

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        width = estimated_column_width(worksheet, column_index)
        worksheet.column_dimensions[letter].width = width
    no_wrap_columns = {
        column_index
        for column_index, cell in enumerate(worksheet[1], start=1)
        if str(cell.value or "").endswith("_ids")
    }
    for row in worksheet.iter_rows(min_row=2):
        worksheet.row_dimensions[row[0].row].height = 18
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=cell.column not in no_wrap_columns,
            )
    return "full"


def estimated_column_width(worksheet, column_index: int) -> int:
    """Estimate a bounded Excel column width from visible cell values."""
    max_length = 8
    for row_index in range(1, min(worksheet.max_row, 200) + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        max_length = max(max_length, len(str(value)))
    return min(max(max_length + 2, 10), 48)


def stable_extra_columns(rows: list[dict[str, Any]], base_columns: list[str]) -> list[str]:
    """Collect non-core columns in first-seen order."""
    seen = set(base_columns)
    extras: list[str] = []
    for row in rows:
        for key in row:
            if str(key).startswith("_"):
                continue
            if key in seen:
                continue
            seen.add(key)
            extras.append(key)
    return extras


def summary_markdown(data: pd.DataFrame) -> str:
    """Render the global summary as readable grouped markdown tables."""
    lines = ["# SQQ summary", ""]
    for title, table in summary_markdown_tables(data):
        if table.empty:
            continue
        lines.extend(["", f"## {title}", "", markdown_table(table).rstrip()])
    return "\n".join(lines).strip() + "\n"


def summary_markdown_tables(data: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Build human-facing markdown summary tables."""
    frame_columns = [
        "frame",
        "time_ps",
        "source",
        "status",
        "n_atoms",
        "n_waters",
        "n_guests",
        "connection_mode",
        "connection_count",
        "hbond_count",
        "oo_connection_count",
        "pair_connection_count",
    ]
    tables: list[tuple[str, pd.DataFrame]] = [
        ("Failures", failure_summary_table(data)),
        ("Frames", summary_simple_table(data, frame_columns)),
        ("Molecules", molecule_summary_table(data)),
        ("Rings", summary_simple_table(data, ["frame", "time_ps", "ring4", "ring5", "ring6", "ring7", "free_ring4", "free_ring5", "free_ring6", "free_ring7"])),
        ("Half Cage", patch_summary_table(data, "half_cage")),
        ("Quasi Cage", patch_summary_table(data, "quasi_cage")),
    ]
    tables.extend(
        [
            ("Cages", cage_summary_table(data)),
            ("Cage Occupancy", cage_occupancy_summary_table(data, markdown_style=True)),
            ("Cage Isomers", cage_isomer_summary_table(data, include_zero_rows=False)),
            ("Order Parameters", order_parameter_summary_table(data)),
            ("Ice", summary_simple_table(data, ["frame", "time_ps", "ice_like_waters", "ice_i_waters", "interfacial_ice_waters"])),
        ]
    )
    return tables


def summary_sheet_tables(data: pd.DataFrame, config: dict[str, Any]) -> dict[str, pd.DataFrame]:
    """Build lightweight main-summary tables using the configured scopes."""
    if is_cpp_mode(config.get("mode", "50")):
        include_zero_isomers = (
            str(config.get("output", {}).get("cage_isomer_rows", "nonzero")).lower()
            == "all"
        )
        tables: dict[str, pd.DataFrame] = {
            "failures": failure_summary_table(data),
            "cage": cage_summary_table(data),
        }
        if has_selected_guests(data):
            tables["cage_occupancy"] = cage_occupancy_summary_table(
                data,
                markdown_style=False,
            )
        tables.update(
            {
                "cage_isomer": cage_isomer_summary_table(
                    data,
                    include_zero_rows=include_zero_isomers,
                ),
                "order_parameter": order_parameter_summary_table(
                    data,
                    config.get("order", {}).get("parameters", ["f3", "f4"]),
                    include_focus=bool(config.get("order", {}).get("focus_waters", [])),
                ),
            }
        )
        return {name: table for name, table in tables.items() if not table.empty}

    ring_sizes = configured_ring_report_sizes(config)
    ring_columns = ["frame", "time_ps"]
    for size in ring_sizes:
        ring_columns.extend([f"ring{size}", f"free_ring{size}"])
    tables: dict[str, pd.DataFrame] = {
        "failures": failure_summary_table(data),
        connection_sheet_name(data): connection_summary_table(data),
        "ring": summary_simple_table(data, ring_columns),
        "half_cage": patch_summary_table(data, "half_cage"),
        "quasi_cage": patch_summary_table(data, "quasi_cage"),
        "cage": cage_summary_table(data),
        "hydrate_cluster": hydrate_cluster_summary_table(data),
        "order_parameter": order_parameter_summary_table(
            data,
            config.get("order", {}).get("parameters", ["f3", "f4"]),
            include_focus=bool(config.get("order", {}).get("focus_waters", [])),
        ),
        "ice": summary_simple_table(data, ["frame", "time_ps", "ice_like_waters", "ice_i_waters", "interfacial_ice_waters"]),
    }
    return {name: table for name, table in tables.items() if not table.empty}


def failure_summary_table(data: pd.DataFrame) -> pd.DataFrame:
    """Return one diagnostic row per failed input frame."""
    if "status" not in data.columns:
        return pd.DataFrame()
    columns = [
        column
        for column in ("frame", "time_ps", "source", "status", "error")
        if column in data.columns
    ]
    failed = data.loc[
        data["status"].astype(str).str.lower().eq("failed"),
        columns,
    ]
    return failed.reset_index(drop=True)


def summary_detail_tables(
    data: pd.DataFrame,
    config: dict[str, Any],
    *,
    raw_rows: list[dict[str, Any]] | None = None,
) -> dict[str, pd.DataFrame]:
    """Build potentially large multi-row tables for summary-detail-csv output."""
    include_zero_isomers = str(config.get("output", {}).get("cage_isomer_rows", "nonzero")).lower() == "all"
    tables: dict[str, pd.DataFrame] = {}
    if output_enabled(config, "summary-detail-csv"):
        tables.update(
            {
                "failures": failure_summary_table(data),
                "cage_occupancy": cage_occupancy_summary_table(
                    data,
                    markdown_style=False,
                ),
                "cage_isomer": cage_isomer_summary_table(
                    data,
                    include_zero_rows=include_zero_isomers,
                ),
                "quasi_cage_isomer": quasi_cage_isomer_summary_table(
                    data,
                    raw_rows=raw_rows,
                ),
            }
        )
    cluster_detail_enabled = output_enabled(config, "cluster-detail")
    if cluster_detail_enabled:
        tables["hydrate_domain"] = hydrate_domain_table(data)
        tables["hydrate_cluster_detail"] = hydrate_cluster_detail_table(data)
    keep_empty: set[str] = set()
    if cluster_detail_enabled and hydrate_cluster_is_enabled(data):
        keep_empty.add("hydrate_domain")
        keep_empty.add("hydrate_cluster_detail")
    return {name: table for name, table in tables.items() if not table.empty or name in keep_empty}


def hydrate_cluster_summary_table(data: pd.DataFrame) -> pd.DataFrame:
    """Build the per-frame hydrate_cluster summary sheet."""
    if not hydrate_cluster_is_enabled(data):
        return pd.DataFrame()
    columns = [
        "frame",
        "time_ps",
        "hydrate_cluster_count",
        "sI_cluster_count",
        "sII_cluster_count",
        "sH_cluster_count",
        "mixed_cluster_count",
        "unclassified_cluster_count",
        "hydrate_domain_count",
        "sI_domain_count",
        "sII_domain_count",
        "sH_domain_count",
        "classified_cage_count",
        "boundary_cage_count",
        "ambiguous_cage_count",
        "unclassified_cage_count",
        "isolated_cage_count",
        "largest_cluster_cage_count",
        "largest_cluster_water_count",
        "cluster_size_distribution",
    ]
    return summary_simple_table(data, columns)


def hydrate_cluster_is_enabled(data: pd.DataFrame) -> bool:
    """Return whether any frame requested hydrate_cluster reporting."""
    if "hydrate_cluster_enabled" not in data.columns:
        return False
    return bool(data["hydrate_cluster_enabled"].astype(str).str.lower().eq("on").any())


def hydrate_cluster_detail_table(data: pd.DataFrame) -> pd.DataFrame:
    """Expand stored per-frame cluster details into one row per cluster."""
    return expanded_hydrate_table(data, "hydrate_cluster_detail")


def hydrate_domain_table(data: pd.DataFrame) -> pd.DataFrame:
    """Expand stored per-frame domain details into one row per domain."""
    return expanded_hydrate_table(data, "hydrate_domain_detail")


def hydrate_motif_table(data: pd.DataFrame) -> pd.DataFrame:
    """Expand stored per-frame motif evidence into one row per motif."""
    return expanded_hydrate_table(data, "hydrate_motif_detail")


def expanded_hydrate_table(data: pd.DataFrame, column: str) -> pd.DataFrame:
    """Expand one stored list of hydrate records into a flat workbook table."""
    schemas = {
        "hydrate_cluster_detail": [
            "cluster_id",
            "hydrate_type",
            "cage_count",
            "water_count",
            "guest_count",
            "empty_cage_count",
            "occupied_cage_count",
            "classified_cage_count",
            "boundary_cage_count",
            "ambiguous_cage_count",
            "unclassified_cage_count",
            "classified_cage_fraction",
            "domain_count",
            "cage_type_counts",
            "cage_composition",
            "boundary_composition",
            "guest_composition",
            "domain_ids",
            "cage_ids",
            "classified_cage_ids",
            "boundary_cage_ids",
            "ambiguous_cage_ids",
            "unclassified_cage_ids",
            "shared_face_count",
        ],
        "hydrate_domain_detail": [
            "domain_id",
            "cluster_id",
            "hydrate_type",
            "status",
            "cage_count",
            "seed_count",
            "seed_cage_count",
            "expanded_cage_count",
            "classified_fraction",
            "water_count",
            "guest_count",
            "external_boundary_contact_count",
            "cage_composition",
            "guest_composition",
            "cage_ids",
            "seed_cage_ids",
            "external_boundary_contact_ids",
        ],
        "hydrate_motif_detail": [
            "motif_id",
            "cluster_id",
            "domain_id",
            "hydrate_type",
            "status",
            "completeness",
            "consistency",
            "confidence",
            "cage_count",
            "core_cage_count",
            "support_cage_count",
            "cage_composition",
            "core_cage_composition",
            "core_cage_ids",
            "motif_cage_ids",
            "internal_shared_face_count",
            "internal_shared_face_ids",
            "classification_method",
        ],
    }
    columns = ["frame", "time_ps", *schemas.get(column, [])]
    if column not in data.columns:
        return pd.DataFrame(columns=columns)
    rows: list[dict[str, Any]] = []
    for _, record in data.iterrows():
        details = record.get(column, [])
        if not isinstance(details, list):
            continue
        for item in details:
            row = {
                "frame": record.get("frame", ""),
                "time_ps": record.get("time_ps", ""),
            }
            row.update(item)
            rows.append(row)
    return pd.DataFrame(rows).reindex(columns=columns)


def connection_sheet_name(data: pd.DataFrame) -> str:
    """Name the connection sheet after the active graph mode."""
    modes = [str(value) for value in data.get("connection_mode", pd.Series(dtype=str)).dropna() if str(value)]
    mode = modes[0] if modes else "connection"
    if mode == "hbond":
        return "hbond"
    if mode == "oo":
        return "oo_connection"
    if mode == "pairs":
        return "pair_connection"
    return "connection"


def connection_summary_table(data: pd.DataFrame) -> pd.DataFrame:
    """Build a per-frame connection and coordination diagnostic table."""
    columns = [
        "frame",
        "time_ps",
        "connection_mode",
        "connection_count",
        "mean_coordination",
        "coordination_0",
        "coordination_1",
        "coordination_2",
        "coordination_3",
        "coordination_4",
        "coordination_gt4",
        "coordination_0_fraction",
        "coordination_1_fraction",
        "coordination_2_fraction",
        "coordination_3_fraction",
        "coordination_4_fraction",
        "coordination_gt4_fraction",
        "degree_le2_fraction",
        "degree4_fraction",
        "over4_fraction",
    ]
    modes = [str(value) for value in data.get("connection_mode", pd.Series(dtype=str)).dropna() if str(value)]
    mode = modes[0] if modes else ""
    mode_column = {
        "hbond": "hbond_count",
        "oo": "oo_connection_count",
        "pairs": "pair_connection_count",
    }.get(mode)
    if mode_column:
        columns.insert(4, mode_column)
    return summary_simple_table(data, columns)


def summary_simple_table(data: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Return a table with only columns present in the run data."""
    existing = [column for column in columns if column in data.columns]
    return data.loc[:, existing].copy() if existing else pd.DataFrame()


def order_parameter_summary_table(
    data: pd.DataFrame,
    parameters: Any | None = None,
    *,
    include_focus: bool | None = None,
) -> pd.DataFrame:
    """Build the per-frame F3/F4/Q_l and hydrate order-parameter table."""
    selected = (
        normalize_order_parameters(parameters)
        if parameters is not None
        else infer_order_parameters_from_data(data)
    )
    if not selected:
        return pd.DataFrame()
    if include_focus is None:
        focus_count_columns = [
            column
            for column in data.columns
            if str(column).endswith("_focus_count")
        ]
        include_focus = any(
            bool((pd.to_numeric(data[column], errors="coerce").fillna(0) > 0).any())
            for column in focus_count_columns
        )
    columns = ["frame", "time_ps"]
    for name in selected:
        if name == "f3":
            columns.extend(["F3_mean", "F3_count"])
        elif name == "f4":
            columns.extend(["F4_mean", "F4_count"])
        elif name.startswith("q"):
            columns.extend([f"{name}_mean", f"{name}_count"])
        elif name == "mcg1":
            columns.append("MCG1_largest_cluster")
        elif name == "mcg3":
            columns.append("MCG3_largest_cluster")
        elif name == "dhop35":
            columns.append("DHOP35_largest_cluster")
        elif name == "dhop30":
            columns.append("DHOP30_largest_cluster")
    if include_focus:
        for name in selected:
            if name == "f3":
                columns.extend(["F3_focus_mean", "F3_focus_count"])
            elif name == "f4":
                columns.extend(["F4_focus_mean", "F4_focus_count"])
            elif name.startswith("q"):
                columns.extend([f"{name}_focus_mean", f"{name}_focus_count"])
    table = data.reindex(columns=columns).copy()
    for column in (
        "MCG1_largest_cluster",
        "MCG3_largest_cluster",
        "DHOP35_largest_cluster",
        "DHOP30_largest_cluster",
    ):
        if column in table:
            table[column] = table[column].where(table[column].notna(), "N/A")
    return table.rename(
        columns={
            "MCG1_largest_cluster": "MCG-1",
            "DHOP35_largest_cluster": "DHOP35",
            "MCG3_largest_cluster": "MCG-3",
            "DHOP30_largest_cluster": "DHOP30",
        }
    )


def infer_order_parameters_from_data(data: pd.DataFrame) -> tuple[str, ...]:
    """Infer legacy table selection when a caller does not provide config."""
    inferred: list[str] = []
    for name, count_column in (("f3", "F3_count"), ("f4", "F4_count")):
        if count_column in data and not data[count_column].replace("", pd.NA).isna().all():
            inferred.append(name)
    inferred.extend(f"q{degree}" for degree in q_degree_from_data(data))
    for name, column in (
        ("mcg1", "MCG1_largest_cluster"),
        ("mcg3", "MCG3_largest_cluster"),
        ("dhop35", "DHOP35_largest_cluster"),
        ("dhop30", "DHOP30_largest_cluster"),
    ):
        if column in data and not data[column].replace("", pd.NA).isna().all():
            inferred.append(name)
    return normalize_order_parameters(inferred or ["none"])


def q_degree_from_data(data: pd.DataFrame) -> list[int]:
    """Infer reported Q_l degree values from summary columns."""
    degree_values: set[int] = set()
    for column in data.columns:
        match = re.fullmatch(r"q(\d+)_(?:mean|count)", str(column))
        if match and not data[column].replace("", pd.NA).isna().all():
            degree_values.add(int(match.group(1)))
    return sorted(degree_values)


def molecule_summary_table(data: pd.DataFrame) -> pd.DataFrame:
    """Build the global molecule-count table using source-file residue order."""
    molecule_columns = [column for column in data.columns if column.startswith("mol_")]
    if not molecule_columns:
        return pd.DataFrame()
    output = pd.DataFrame({"frame": data["frame"], "time_ps": data["time_ps"]})
    for column in molecule_columns:
        output[column.removeprefix("mol_")] = [count_cell(value) for value in data[column]]
    return output


def patch_summary_table(data: pd.DataFrame, prefix: str) -> pd.DataFrame:
    """Build a plotting-friendly half_cage or quasi_cage table."""
    columns = [column for column in data.columns if column.startswith(f"{prefix}_") and column not in {f"{prefix}_total", f"{prefix}_breakdown"}]
    if not columns:
        return pd.DataFrame()
    sorted_columns = sorted(columns)
    labels = [patch_summary_label(column.removeprefix(f"{prefix}_"), prefix) for column in sorted_columns]
    output_labels = sorted(set(labels))
    rows: list[dict[str, Any]] = []
    for _, record in data.iterrows():
        row: dict[str, Any] = {
            "frame": record.get("frame", ""),
            "time_ps": record.get("time_ps", ""),
        }
        total = 0
        for column, label in zip(sorted_columns, labels):
            count = count_cell(record.get(column, 0))
            row[label] = row.get(label, 0) + count
            total += count
        row["total"] = total
        rows.append(row)
    return pd.DataFrame(rows).reindex(columns=["frame", "time_ps", *output_labels, "total"])


def patch_summary_label(patch_type: str, prefix: str) -> str:
    """Return the workbook label for an open-patch type."""
    if prefix != "quasi_cage":
        return patch_type
    return patch_composition_label(patch_type)


def patch_composition_label(patch_type: str) -> str:
    """Return a patch label without internal prefix or ring-sequence isomer marks."""
    return patch_display_label(patch_type).translate(SUBSCRIPT_DIGIT_DELETE)


def summary_cage_types_from_data(data: pd.DataFrame) -> list[str]:
    """Collect exact cage report types, including requested zero-count types."""
    requested: set[str] = set()
    report_all = False
    for value in data.get("cage_report_types", pd.Series(dtype=str)).dropna():
        marker = str(value).strip()
        if not marker:
            continue
        if marker.lower() == "all":
            report_all = True
            continue
        requested.update(item for item in marker.split(";") if item)
    if requested and not report_all:
        return [item for item in ordered_cage_types(requested) if item in requested]

    detected: set[str] = set()
    for column in data.columns:
        if not column.startswith("cage_"):
            continue
        label = column.removeprefix("cage_")
        if label in {"empty", "occupied", "total", "report_types"} or "_" in label:
            continue
        values = pd.to_numeric(data[column], errors="coerce").fillna(0)
        if bool((values > 0).any()):
            detected.add(label)
    return [item for item in ordered_cage_types(detected) if item in detected]


def cage_summary_table(data: pd.DataFrame) -> pd.DataFrame:
    """Build the global cage-count table with superscript cage headers."""
    cage_types = summary_cage_types_from_data(data)
    output = pd.DataFrame({"frame": data["frame"], "time_ps": data["time_ps"]})
    for cage_type in cage_types:
        output[cage_display_label(cage_type)] = [count_cell(value) for value in data.get(f"cage_{cage_type}", pd.Series([0] * len(data)))]
    output["total"] = output[[cage_display_label(cage_type) for cage_type in cage_types]].sum(axis=1)
    return output


def cage_occupancy_summary_table(data: pd.DataFrame, markdown_style: bool) -> pd.DataFrame:
    """Build global cage occupancy rows, optionally with tree markers."""
    cage_types = summary_cage_types_from_data(data)
    guest_labels = global_guest_labels(data, cage_types)
    rows: list[dict[str, Any]] = []
    for _, record in data.iterrows():
        child_labels = row_guest_labels(record, guest_labels, cage_types)
        if sum(cage_count(record, cage_type, "multi") for cage_type in cage_types) > 0:
            child_labels.append("multi")
        for label in ["empty", "occupied", *child_labels]:
            counts = [cage_count(record, cage_type, label) for cage_type in cage_types]
            branch = ""
            display_label = label
            if label not in {"empty", "occupied"} and markdown_style:
                branch = "└" if label == child_labels[-1] else "├"
                display_label = f"{branch} {label}"
            row: dict[str, Any] = {
                "frame": record.get("frame", ""),
                "time_ps": record.get("time_ps", ""),
                "occupancy": display_label,
            }
            if not markdown_style and label not in {"empty", "occupied"}:
                row["level"] = "detail"
            elif not markdown_style:
                row["level"] = "class"
            for cage_type, count in zip(cage_types, counts):
                row[cage_display_label(cage_type)] = f"{branch} {count}" if branch else count
            row["total"] = f"{branch} {sum(counts)}" if branch else sum(counts)
            rows.append(row)
        total_counts = [cage_count(record, cage_type, "empty") + cage_count(record, cage_type, "occupied") for cage_type in cage_types]
        row = {"frame": record.get("frame", ""), "time_ps": record.get("time_ps", ""), "occupancy": "total"}
        if not markdown_style:
            row["level"] = "total"
        for cage_type, count in zip(cage_types, total_counts):
            row[cage_display_label(cage_type)] = count
        row["total"] = sum(total_counts)
        rows.append(row)
    return pd.DataFrame(rows)


def global_guest_labels(data: pd.DataFrame, cage_types: list[str]) -> list[str]:
    """Collect guest residue labels from occupancy columns in first-seen order."""
    labels: list[str] = []
    for value in data.get("guest_order", pd.Series(dtype=str)):
        if pd.isna(value) or value == "":
            continue
        for label in str(value).split(";"):
            if label and label not in labels:
                labels.append(label)
    for column in data.columns:
        label = cage_occupancy_label_from_column(column, cage_types)
        if label and label not in {"empty", "occupied", "multi"} and label not in labels:
            labels.append(label)
    return labels


def row_guest_labels(record: pd.Series, labels: list[str], cage_types: list[str]) -> list[str]:
    """Return guest labels that occur in this frame, preserving global order."""
    result: list[str] = []
    for label in labels:
        if any(cage_count(record, cage_type, label) for cage_type in cage_types):
            result.append(label)
    return result


def cage_occupancy_label_from_column(column: str, cage_types: list[str]) -> str | None:
    """Extract MET from cage_512_MET style columns."""
    if not column.startswith("cage_"):
        return None
    label = None
    for cage_type in sorted(cage_types, key=len, reverse=True):
        prefix = f"cage_{cage_type}_"
        if column.startswith(prefix):
            label = column.removeprefix(prefix)
            break
    if label is None:
        return None
    if label.startswith("isomer") or label == "isomers":
        return None
    return label


def cage_count(record: pd.Series, cage_type: str, label: str) -> int:
    """Read one cage occupancy count from a summary row."""
    return count_cell(record.get(f"cage_{cage_type}_{label}", 0))


def cage_isomer_summary_table(data: pd.DataFrame, include_zero_rows: bool) -> pd.DataFrame:
    """Build global cage-isomer rows with separate columns."""
    cage_types = summary_cage_types_from_data(data)
    labels = global_cage_isomer_labels(data, cage_types)
    if not labels:
        return pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for _, record in data.iterrows():
        for label in labels:
            row: dict[str, Any] = {
                "frame": record.get("frame", ""),
                "time_ps": record.get("time_ps", ""),
                "isomer": label,
            }
            total = 0
            for cage_type in cage_types:
                column = f"cage_{cage_type}_isomer_{label}"
                if column not in data.columns:
                    row[cage_display_label(cage_type)] = "-"
                    continue
                count = count_cell(record.get(column, 0))
                row[cage_display_label(cage_type)] = count
                total += count
            if total == 0 and not include_zero_rows:
                continue
            row["total"] = total
            rows.append(row)
        total_row: dict[str, Any] = {"frame": record.get("frame", ""), "time_ps": record.get("time_ps", ""), "isomer": "total"}
        total = 0
        for cage_type in cage_types:
            count = count_cell(record.get(f"cage_{cage_type}", 0))
            total_row[cage_display_label(cage_type)] = count
            total += count
        total_row["total"] = total
        rows.append(total_row)
    return pd.DataFrame(rows)


def global_cage_isomer_labels(data: pd.DataFrame, cage_types: list[str]) -> list[str]:
    """Collect cage-isomer labels in canonical cage-type order."""
    labels: list[str] = []
    for cage_type in cage_types:
        prefix = f"cage_{cage_type}_isomer_"
        for column in data.columns:
            if column.startswith(prefix):
                label = column.removeprefix(prefix)
                if label not in labels:
                    labels.append(label)
    return labels


def quasi_cage_isomer_summary_table(
    data: pd.DataFrame,
    *,
    raw_rows: list[dict[str, Any]] | None = None,
) -> pd.DataFrame:
    """Build long-form quasi-cage isomer rows for CSV detail output."""
    output_columns = ["frame", "time_ps", "quasi_cage_type", "isomer", "count"]
    rows: list[dict[str, Any]] = []
    if raw_rows is not None:
        for record in raw_rows:
            details = record.get(QUASI_ISOMER_DETAIL_KEY, ())
            for isomer, count in details:
                count = count_cell(count)
                if count:
                    rows.append(
                        {
                            "frame": record.get("frame", ""),
                            "time_ps": record.get("time_ps", ""),
                            "quasi_cage_type": patch_composition_label(isomer),
                            "isomer": patch_display_label(isomer),
                            "count": count,
                        }
                    )
        return pd.DataFrame(rows, columns=output_columns)

    # Support legacy wide-row callers.
    columns = [
        column
        for column in data.columns
        if column.startswith("quasi_cage_") and column not in {"quasi_cage_total", "quasi_cage_breakdown"}
    ]
    for _, record in data.iterrows():
        for column in sorted(columns):
            count = count_cell(record.get(column, 0))
            if count == 0:
                continue
            isomer = column.removeprefix("quasi_cage_")
            rows.append(
                {
                    "frame": record.get("frame", ""),
                    "time_ps": record.get("time_ps", ""),
                    "quasi_cage_type": patch_composition_label(isomer),
                    "isomer": patch_display_label(isomer),
                    "count": count,
                }
            )
    return pd.DataFrame(rows, columns=output_columns)


def count_cell(value: Any) -> int:
    """Normalize missing numeric count cells to zero."""
    if value is None or value == "":
        return 0
    try:
        if pd.isna(value):
            return 0
    except TypeError:
        pass
    return int(value)


def markdown_table(data: pd.DataFrame) -> str:
    """Render a pandas DataFrame as a simple GitHub-style table."""
    headers = [str(col) for col in data.columns]
    body = [[format_summary_cell(value) for value in row] for row in data.itertuples(index=False, name=None)]
    widths = [len(header) for header in headers]
    for row in body:
        for idx, value in enumerate(row):
            widths[idx] = max(widths[idx], len(value))
    lines = [
        "| " + " | ".join(header.ljust(widths[idx]) for idx, header in enumerate(headers)) + " |",
        "| " + " | ".join("-" * widths[idx] for idx in range(len(headers))) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(value.ljust(widths[idx]) for idx, value in enumerate(row)) + " |")
    return "\n".join(lines) + "\n"


def format_summary_cell(value: Any) -> str:
    """Format markdown cells while keeping count columns readable."""
    if isinstance(value, (list, tuple, dict)):
        return repr(value)
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def excel_scalar(value: Any) -> Any:
    """Convert containers into readable scalar values for XLSX cells."""
    if isinstance(value, (list, tuple, set)):
        return ",".join(str(item) for item in value)
    if isinstance(value, dict):
        return repr(value)
    return value


def flatten_config(config: dict[str, Any], prefix: str = "") -> list[dict[str, str]]:
    """Flatten nested config keys for the main summary config table."""
    rows = []
    for key, value in config.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(flatten_config(value, name))
        else:
            rows.append({"parameter": name, "value": repr(value)})
    return rows
